"""Routen /api/nachkalkulation und /api/importe (PLAN §4, §7 Phase 4, §8).

Zwei Dinge stehen hier im Vordergrund:

* **Die Trennung von Projektsicht und Finanzsicht** (PLAN §4). ``team`` sieht Projekte, aber
  keine Margen. ``buchhaltung`` darf importieren, sieht die Nachkalkulation aber nicht. Das ist
  keine Kleinigkeit: die Rollentabelle in PLAN §4 wäre ohne diese Trennung sinnlos.
* **Vorschau und Übernahme hängen zusammen.** Ändert sich die Datei dazwischen, wird der Lauf
  abgewiesen – sonst wird etwas anderes geschrieben, als auf dem Schirm stand.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from sqlalchemy import select

from app.datenbank import schreib_sitzung
from app.modelle import (
    Firma,
    IstKosten,
    Kunde,
    Projekt,
    SollKalkulation,
    Stuecklistenposition,
    Stunden,
)
from tests.conftest_auth import anmelden

KOPF = (
    "Belegdatum;Konto;Kontobezeichnung;Buchungstext;Belegfeld 1;Umsatz;Soll/Haben-Kennzeichen;KOST2"
)


@pytest.fixture
def bestand(gesäte_db) -> None:
    """Zwei Projekte: 26001 mit voller Nachkalkulation, 26002 ohne Kalkulationsblatt."""
    with schreib_sitzung() as sitzung:
        firma_id = sitzung.scalar(select(Firma.id).order_by(Firma.id).limit(1))
        kunde = Kunde(kunden_nr=19001, name="Route GmbH", ort="Weiden", typ="b2b")
        sitzung.add(kunde)
        sitzung.flush()

        eins = Projekt(
            projekt_nr=26001,
            firma_id=firma_id,
            kunde_id=kunde.id,
            status="abgeschlossen",
            ab_wert_netto=10000000,
            pl_name="Stefan",
        )
        zwei = Projekt(
            projekt_nr=26002,
            firma_id=firma_id,
            kunde_id=kunde.id,
            status="in_bau",
            ab_wert_netto=5000000,
            pl_name="Günther",
        )
        sitzung.add_all([eins, zwei])
        sitzung.flush()

        sitzung.add(
            SollKalkulation(
                projekt_id=eins.id,
                material_soll=6000000,
                dl_soll=500000,
                stunden_soll=Decimal("100.00"),
                marge_soll=180,
            )
        )
        sitzung.add(
            IstKosten(
                projekt_id=eins.id,
                quelle="datev",
                monat="2026-07",
                betrag=7000000,
                referenz="3400 Wareneingang",
            )
        )
        sitzung.add(
            Stunden(
                projekt_id=eins.id,
                monat="2026-07",
                mitarbeiter="Wilhelm, Sven",
                stunden=Decimal("95.00"),
                satz=8500,
            )
        )
        sitzung.add(
            IstKosten(
                projekt_id=eins.id,
                quelle="timetac",
                monat="2026-07",
                betrag=807500,
                referenz="Arbeitsstunden (TimeTac)",
            )
        )
        sitzung.add(
            Stuecklistenposition(
                projekt_id=eins.id,
                artikel_nr="MOD-450",
                bezeichnung="Modul 450 Wp",
                menge_soll=Decimal("88.000"),
                menge_ist=Decimal("88.000"),
                ek_preis=9240,
                quelle="projektbestellt",
            )
        )
        sitzung.add(
            Stuecklistenposition(
                projekt_id=zwei.id,
                bezeichnung="Schiene",
                menge_soll=Decimal("36.000"),
                ek_preis=2190,
                quelle="lager",
            )
        )


@pytest.fixture
def admin(client, nutzer_erzeugen, bestand):
    nutzer_erzeugen("chef@ip3-energie.de", "admin")
    return anmelden(client, "chef@ip3-energie.de")


@pytest.fixture
def buchhaltung(client, nutzer_erzeugen, bestand):
    nutzer_erzeugen("buha@ip3-energie.de", "buchhaltung")
    return anmelden(client, "buha@ip3-energie.de")


@pytest.fixture
def team(client, nutzer_erzeugen, bestand):
    nutzer_erzeugen("team@ip3-energie.de", "team")
    return anmelden(client, "team@ip3-energie.de")


# ---------------------------------------------------------------------------
# Rechte
# ---------------------------------------------------------------------------


def test_team_sieht_projekte_aber_keine_margen(team) -> None:
    """PLAN §4: Finanzsichtbarkeit ist bewusst von der Projektsicht getrennt."""
    assert team.client.get("/api/projekte").status_code == 200
    assert team.client.get("/api/nachkalkulation").status_code == 403
    assert team.client.get("/api/nachkalkulation/26001").status_code == 403


def test_buchhaltung_darf_importieren_aber_keine_margen_sehen(buchhaltung) -> None:
    assert buchhaltung.client.get("/api/nachkalkulation").status_code == 403
    assert buchhaltung.client.get("/api/importe/laeufe").status_code == 200


def test_team_darf_nicht_importieren(team) -> None:
    assert team.client.get("/api/importe/laeufe").status_code == 403
    assert (
        team.schreiben("POST", "/api/importe/datev/uebernehmen", json={"kennung": "x"}).status_code
        == 403
    )


def test_ohne_anmeldung_kein_zugriff(client) -> None:
    assert client.get("/api/nachkalkulation").status_code == 401
    assert client.get("/api/importe/laeufe").status_code == 401


# ---------------------------------------------------------------------------
# Übersicht
# ---------------------------------------------------------------------------


def test_uebersicht_liefert_zahlen_und_summen(admin) -> None:
    inhalt = admin.client.get("/api/nachkalkulation").json()

    assert inhalt["anzahl"] == 2
    eins = next(p for p in inhalt["projekte"] if p["projekt_nr"] == 26001)
    assert eins["erloes_netto"] == 10000000
    assert eins["ist_gesamt"] == 7807500
    assert eins["marge_netto"] == 2192500
    assert eins["marge_promille"] == 219
    assert eins["ampel"] == "im_soll"
    assert eins["soll_gesamt"] == 6500000
    assert eins["stunden_ist"] == "95.00"

    assert inhalt["erloes_netto"] == 15000000
    assert inhalt["ohne_kalkulation"] == 1
    assert inhalt["ampel_gelb_promille"] == 50


def test_uebersicht_sortiert_die_schwaechste_marge_nach_oben(admin) -> None:
    inhalt = admin.client.get("/api/nachkalkulation").json()
    assert inhalt["projekte"][0]["projekt_nr"] == 26001


def test_sortierung_nach_projektnummer(admin) -> None:
    inhalt = admin.client.get("/api/nachkalkulation?sortierung=projekt_nr").json()
    assert [p["projekt_nr"] for p in inhalt["projekte"]] == [26001, 26002]


def test_filter_nach_status_und_projektleiter(admin) -> None:
    nur_bau = admin.client.get("/api/nachkalkulation?status=in_bau").json()
    assert [p["projekt_nr"] for p in nur_bau["projekte"]] == [26002]

    nur_stefan = admin.client.get("/api/nachkalkulation?projektleiter=Stefan").json()
    assert [p["projekt_nr"] for p in nur_stefan["projekte"]] == [26001]


def test_filter_nur_mit_hinweis(admin) -> None:
    inhalt = admin.client.get("/api/nachkalkulation?nur_mit_hinweis=true").json()
    assert [p["projekt_nr"] for p in inhalt["projekte"]] == [26002]
    assert inhalt["anzahl"] == 2, "die Summen zählen weiter alle Projekte"


def test_unsinniger_filterwert_ergibt_eine_meldung(admin) -> None:
    """Ein Tippfehler darf nicht wie „keine Projekte" aussehen."""
    assert admin.client.get("/api/nachkalkulation?status=gibtsnicht").status_code == 422


# ---------------------------------------------------------------------------
# Projektansicht
# ---------------------------------------------------------------------------


def test_projektansicht_mit_aufgliederung(admin) -> None:
    inhalt = admin.client.get("/api/nachkalkulation/26001").json()

    assert inhalt["projekt"]["marge_promille"] == 219
    assert inhalt["stunden"] == [
        {
            "monat": "2026-07",
            "mitarbeiter": "Wilhelm, Sven",
            "stunden": "95.00",
            "satz": 8500,
            "betrag": 807500,
        }
    ]
    assert inhalt["stueckliste"][0]["artikel_nr"] == "MOD-450"
    assert inhalt["stueckliste"][0]["bewertet_betrag"] is None, "projektbestellt bleibt unbewertet"


def test_unbekanntes_projekt_ergibt_404_mit_naechstem_schritt(admin) -> None:
    antwort = admin.client.get("/api/nachkalkulation/99999")
    assert antwort.status_code == 404
    assert "Projektnummer prüfen" in antwort.json()["naechster_schritt"]


def test_offene_mengen_liste(admin) -> None:
    inhalt = admin.client.get("/api/nachkalkulation/mengen-ist-offen").json()
    assert [e["projekt_nr"] for e in inhalt] == [26002]
    assert inhalt[0]["offen"] == 1


class TestScope:
    """Ein Projektleiter mit Scope ``eigene`` sieht auch nur seine Margen.

    Die Rolle wird eigens angelegt: ``team`` trägt ``projekte.lesen`` mit Scope ``alle``, und
    eine zusätzliche Berechtigung mit ``eigene`` würde daran nichts ändern – die weitere gewinnt.
    """

    @pytest.fixture
    def projektleiter(self, client, nutzer_erzeugen, bestand):
        from app.modelle import Berechtigung, Rolle, User

        nutzer_id = nutzer_erzeugen("pl@ip3-energie.de", "team")
        with schreib_sitzung() as sitzung:
            rolle = Rolle(name="pl-nachkalkulation", beschreibung="Nur eigene Projekte")
            sitzung.add(rolle)
            eigene = Berechtigung(
                schluessel="projekte.lesen", scope="eigene", beschreibung="nur eigene"
            )
            sitzung.add(eigene)
            rolle.berechtigungen.append(eigene)
            recht = sitzung.scalar(
                select(Berechtigung).where(
                    Berechtigung.schluessel == "nachkalkulation.lesen",
                    Berechtigung.scope.is_(None),
                )
            )
            assert recht is not None
            rolle.berechtigungen.append(recht)

            nutzer = sitzung.get(User, nutzer_id)
            nutzer.rollen.clear()
            nutzer.rollen.append(rolle)
            sitzung.flush()
            # Nur 26002 gehört diesem Konto.
            projekt = sitzung.scalar(select(Projekt).where(Projekt.projekt_nr == 26002))
            projekt.pl_user_id = nutzer_id
        return anmelden(client, "pl@ip3-energie.de")

    def test_scope_beschraenkt_die_uebersicht(self, projektleiter) -> None:
        inhalt = projektleiter.client.get("/api/nachkalkulation").json()
        assert [p["projekt_nr"] for p in inhalt["projekte"]] == [26002]
        assert inhalt["erloes_netto"] == 5000000, "auch die Summen folgen dem Scope"

    def test_scope_beschraenkt_auch_die_projektansicht(self, projektleiter) -> None:
        assert projektleiter.client.get("/api/nachkalkulation/26002").status_code == 200
        assert projektleiter.client.get("/api/nachkalkulation/26001").status_code == 404

    def test_scope_beschraenkt_die_liste_der_offenen_mengen(self, projektleiter) -> None:
        inhalt = projektleiter.client.get("/api/nachkalkulation/mengen-ist-offen").json()
        assert [e["projekt_nr"] for e in inhalt] == [26002]


# ---------------------------------------------------------------------------
# Importe
# ---------------------------------------------------------------------------


class TestDatevImport:
    @pytest.fixture
    def datev_ordner(self, tmp_path: Path, test_einstellungen) -> Path:
        ordner = tmp_path / "02_DATEV"
        ordner.mkdir()
        test_einstellungen.pfade.datev = ordner
        return ordner

    def datei_schreiben(self, ordner: Path, zeilen: list[str], name="kostentraeger_2026-08.csv"):
        (ordner / name).write_text("\n".join([KOPF, *zeilen]) + "\n", encoding="utf-8")
        return ordner / name

    def test_vorschau_schreibt_nichts(self, buchhaltung, datev_ordner: Path) -> None:
        self.datei_schreiben(
            datev_ordner,
            ["05.08.2026;3400;Wareneingang;Module;RE-1;1.000,00;S;26001"],
        )
        antwort = buchhaltung.client.get("/api/importe/datev/vorschau")
        assert antwort.status_code == 200
        inhalt = antwort.json()
        assert inhalt["zeitraum"] == "2026-08"
        assert inhalt["kontrollsummen"]["summe_cent"] == 100000
        assert buchhaltung.client.get("/api/importe/laeufe").json() == []

    def test_uebernahme_mit_der_kennung_aus_der_vorschau(
        self, buchhaltung, datev_ordner: Path
    ) -> None:
        self.datei_schreiben(
            datev_ordner,
            ["05.08.2026;3400;Wareneingang;Module;RE-1;1.000,00;S;26001"],
        )
        vorschau = buchhaltung.client.get("/api/importe/datev/vorschau").json()
        antwort = buchhaltung.schreiben(
            "POST",
            "/api/importe/datev/uebernehmen",
            json={"kennung": vorschau["kennung"]},
        )
        assert antwort.status_code == 200, antwort.text
        inhalt = antwort.json()
        assert inhalt["zeitraum"] == "2026-08"
        assert inhalt["ergebnis"]["summe_cent"] == 100000
        assert "übernommen" in inhalt["meldung"]

    def test_geaenderte_datei_wird_abgewiesen(self, buchhaltung, datev_ordner: Path) -> None:
        """Sonst wird etwas anderes geschrieben, als auf dem Schirm stand."""
        self.datei_schreiben(
            datev_ordner,
            ["05.08.2026;3400;Wareneingang;Module;RE-1;1.000,00;S;26001"],
        )
        vorschau = buchhaltung.client.get("/api/importe/datev/vorschau").json()
        self.datei_schreiben(
            datev_ordner,
            ["05.08.2026;3400;Wareneingang;Module;RE-1;9.999,00;S;26001"],
        )
        antwort = buchhaltung.schreiben(
            "POST",
            "/api/importe/datev/uebernehmen",
            json={"kennung": vorschau["kennung"]},
        )
        assert antwort.status_code == 409
        assert antwort.json()["code"] == "import_datei_geaendert"

    def test_fehlender_ordner_nennt_den_naechsten_schritt(
        self, buchhaltung, test_einstellungen
    ) -> None:
        test_einstellungen.pfade.datev = None
        antwort = buchhaltung.client.get("/api/importe/datev/vorschau")
        assert antwort.status_code == 409
        assert "config.toml" in antwort.json()["naechster_schritt"]

    def test_fehlende_datei_nennt_das_dateimuster(self, buchhaltung, datev_ordner: Path) -> None:
        antwort = buchhaltung.client.get("/api/importe/datev/vorschau")
        assert antwort.status_code == 409
        assert "kostentraeger_JJJJ-MM.csv" in antwort.json()["naechster_schritt"]

    def test_lauf_steht_anschliessend_im_protokoll(self, buchhaltung, datev_ordner: Path) -> None:
        self.datei_schreiben(
            datev_ordner,
            ["05.08.2026;3400;Wareneingang;Module;RE-1;1.000,00;S;26001"],
        )
        vorschau = buchhaltung.client.get("/api/importe/datev/vorschau").json()
        buchhaltung.schreiben(
            "POST", "/api/importe/datev/uebernehmen", json={"kennung": vorschau["kennung"]}
        )
        laeufe = buchhaltung.client.get("/api/importe/laeufe").json()
        assert len(laeufe) == 1
        assert laeufe[0]["quelle"] == "datev"
        assert laeufe[0]["zeitraum"] == "2026-08"
        assert laeufe[0]["status"] == "erfolg"


class TestKalkulationImport:
    @pytest.fixture
    def kalkulationsordner(self, tmp_path: Path, test_einstellungen) -> Path:
        from app.importe.kalkulationsblatt import vorlage_erzeugen

        ordner = tmp_path / "03_Kalkulation"
        ordner.mkdir()
        test_einstellungen.pfade.kalkulation = ordner

        from openpyxl import load_workbook

        pfad = vorlage_erzeugen(ordner / "26001_Route.xlsx")
        mappe = load_workbook(pfad)
        blatt = mappe["EXPORT"]
        blatt["B6"] = 26001
        blatt["B7"] = 60000
        blatt["B8"] = 5000
        blatt["B9"] = 100
        blatt["B10"] = 18
        blatt["A14"], blatt["B14"], blatt["C14"] = "MOD-450", "Modul 450 Wp", 88
        blatt["D14"], blatt["E14"], blatt["F14"] = 92.40, "projektbestellt", "pv"
        mappe.save(pfad)
        mappe.close()
        return ordner

    def test_vorschau_zaehlt_die_blaetter(self, buchhaltung, kalkulationsordner: Path) -> None:
        inhalt = buchhaltung.client.get("/api/importe/kalkulation/vorschau").json()
        assert inhalt["kontrollsummen"]["blaetter"] == 1
        assert inhalt["kontrollsummen"]["positionen"] == 1
        assert inhalt["dateien"] == ["26001_Route.xlsx"]

    def test_uebernahme_schreibt_die_sollwerte(self, buchhaltung, kalkulationsordner: Path) -> None:
        vorschau = buchhaltung.client.get("/api/importe/kalkulation/vorschau").json()
        antwort = buchhaltung.schreiben(
            "POST",
            "/api/importe/kalkulation/uebernehmen",
            json={"kennung": vorschau["kennung"]},
        )
        assert antwort.status_code == 200, antwort.text
        assert antwort.json()["ergebnis"]["uebernommen"] == 1

        from app.datenbank import lese_sitzung

        with lese_sitzung() as sitzung:
            projekt = sitzung.scalar(select(Projekt).where(Projekt.projekt_nr == 26001))
            soll = sitzung.get(SollKalkulation, projekt.id)
            assert soll.material_soll == 6000000
            assert soll.marge_soll == 180

    def test_kaputtes_blatt_haelt_die_uebrigen_nicht_auf(
        self, buchhaltung, kalkulationsordner: Path
    ) -> None:
        from openpyxl import Workbook

        mappe = Workbook()
        mappe.active.title = "Kalkulation"
        mappe.save(kalkulationsordner / "26002_Ohne_Export.xlsx")
        mappe.close()

        inhalt = buchhaltung.client.get("/api/importe/kalkulation/vorschau").json()
        assert inhalt["kontrollsummen"]["blaetter"] == 1, "das heile Blatt kommt durch"
        assert any("26002" in b["datei"] for b in inhalt["befunde"])
