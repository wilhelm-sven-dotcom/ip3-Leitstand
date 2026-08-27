"""Lesen der beiden Excel-Bestandsdateien (PLAN §9).

Beide Dateien werden **ausschließlich gelesen**. Der Leser urteilt nicht und wirft nichts weg:
jeder Wert, den er nicht sicher deuten kann, wird zu einem :class:`Befund` mit Zeile, Spalte und
Originalinhalt. Die Befunde stehen später im Importprotokoll und in der Zuordnungsmaske. Ein
Abbruch mitten in der Datei wäre das Schlechteste – dann sieht niemand, wie viel in Ordnung war.

Die Kontrollsummen rechnet der Leser selbst über die Datenzeilen. Die Summenzellen der Dateien
sind nachweislich falsch (siehe :data:`SUMMENFEHLER`), taugen also nicht als Vorgabe. Beide
Werte stehen im Protokoll, damit die Abweichung erklärt ist und nicht wie ein Importfehler
aussieht.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.fehler import FachFehler
from app.geld import euro_nach_cent
from app.migration.vokabular import Rechnungsart, kunde_und_ort, rechnungsart_lesen

# ---------------------------------------------------------------------------
# Auftragsliste: Blatt 'Et-Einnahmen'
# ---------------------------------------------------------------------------

BLATT_AUFTRAEGE = "Et-Einnahmen"
ERSTE_DATENZEILE = 8
SPALTE_TEXT = "A"
SPALTE_BETRAG = "B"
SPALTE_GESTELLT = "D"

# Je Monat ein Spaltenpaar: in der Markerspalte steht das 'x', die Spalte daneben trägt nur die
# Formel ``=WENN(<Marker>="x";$B<zeile>;"")``. Gelesen wird der Marker, nicht das Ergebnis.
# November bricht das Muster (AJ statt AK) – die Datei ist über Jahre gewachsen.
MARKERSPALTEN: tuple[str, ...] = ("G", "J", "M", "P", "S", "V", "Y", "AB", "AE", "AH", "AJ")

# ---------------------------------------------------------------------------
# Teamliste: Blatt 'Übersicht Projekte'
# ---------------------------------------------------------------------------

BLATT_PROJEKTE = "Übersicht Projekte"
SPALTE_KUNDE = "B"

# Stammdaten des Projekts.
SPALTE_PV_KWP = "C"
SPALTE_WR = "D"
SPALTE_SPEICHER = "E"
SPALTE_STORAGE = "F"
SPALTE_LADESTATION = "G"
SPALTE_AUFTRAG_VOM = "H"
SPALTE_AB_WERT = "I"
SPALTE_PL = "R"
SPALTE_BEMERKUNG = "BG"

# Nachkalkulations-Altwerte. Werden als Notiz übernommen, nicht als Ist-Kosten (PLAN §9).
NACHKALKULATION: dict[str, str] = {
    "J": "Status",
    "K": "Bemerkung",
    "L": "Netto-Ausgaben",
    "M": "Nachkalkulation €",
    "N": "GWL/Risiko/Verwaltung",
    "O": "Vertrieb",
    "P": "Gewinn %",
}

# Terminblock. Jede Spalte bekommt einen eigenen Meilenstein-Typ, weil UNIQUE(projekt_id, typ)
# gilt und die acht Spalten sonst auf zwei Zeilen zusammenfielen (siehe Migration 0003).
TERMINSPALTEN: dict[str, str] = {
    "AC": "montage_uk",
    "AD": "montage_elektro",
    "AE": "zaehlerschrank",
    "AF": "lieferung_uk",
    "AG": "lieferung_wr_pv",
    "AH": "lieferung_wr_speicher",
    "AI": "lieferung_speicher",
    "AJ": "lieferung_wallbox",
}

# Statusblock.
STATUSSPALTEN: dict[str, str] = {
    "AM": "uebergabetermin",
    "AN": "freigabe_planung",
    "AO": "plan_erstellt",
    "AP": "anmeldung_nb",
    "AQ": "mastr",
    "AR": "fertigmeldung",
    "AS": "zaehler",
    "AT": "abnahme",
}

# Spalte AK trägt die Überschrift 'Module reserviert [Stück]', enthält aber Bruchzahlen
# (24,888…) und #VALUE!-Fehler: eine Rechenspalte, kein Termin. Sie wird nicht übernommen.
SPALTE_MODULE_RESERVIERT = "AK"

# Der Vorplanungsblock (T–AA) führt Kalenderwochen für Schritte, die der Termin- und
# Statusblock schon abdeckt, und ist mit 14 gefüllten Zellen auf 530 Zeilen praktisch leer.
# Abweichend von PLAN §9 entstehen daraus keine Meilensteine – acht weitere Typen für
# Doppelungen ohne Daten wären Schemaballast. Die Werte gehen als Notiz an das Projekt und
# erzeugen je Zelle einen Befund, damit im Protokoll steht, wo sie geblieben sind.
VORPLANUNGSSPALTEN: dict[str, str] = {
    "T": "1. Übernahme",
    "U": "2. Planung",
    "V": "3. Bestellung",
    "W": "4. Lieferung",
    "X": "5. Dachmontage",
    "Y": "6. AC-Montage",
    "Z": "7. Doku",
    "AA": "8. Abnahme",
}

# Abschlagskreuze der Buchhaltungsspalten. Sie sagen, welche Abschläge gestellt sind, und
# dienen der Gegenprobe zur Auftragsliste.
ABSCHLAGSSPALTEN: dict[str, tuple[str, int]] = {
    "AV": ("pv", 1),
    "AW": ("pv", 2),
    "AX": ("pv", 3),
    "AY": ("pv", 4),
    "AZ": ("speicher", 1),
    "BA": ("speicher", 2),
    "BB": ("speicher", 3),
    "BC": ("speicher", 4),
}

# Falsche Summenformeln in den Quelldateien, am Original nachgemessen. Der Importer nennt sie
# im Protokoll, damit die Abweichung zu den gewohnten Zahlen erklärt ist.
SUMMENFEHLER: tuple[tuple[str, str, str], ...] = (
    (
        "Teambesprechung_NEU",
        "I7",
        "SUMME(I24:I527) statt über alle Datenzeilen – übergeht 29 Projekte",
    ),
    ("Teambesprechung_NEU", "C6", "SUMME(C24:C527), derselbe Bereichsfehler"),
    (
        "Offene_Auftraege",
        "Z5",
        "SUMME(Z8:AC3243) summiert ein Rechteck und zählt den August in den Juli",
    ),
)

MARKER_JA = ("x", "X", "✓")
MARKER_NEIN = ("-", "–", "—")
MARKER_OFFEN = ("o", "O")
_KALENDERWOCHE = re.compile(r"^\s*(\d{1,2})\s*/\s*(\d{2,4})\s*$")

# Dass eine Anlage auf einer Freifläche steht, sagt kein Feld der Teamliste – es steht allenfalls
# im Klartext des Namens. Der Ausdruck ist bewusst kurz gehalten: am echten Bestand geprüft
# findet er genau eine Zeile („Thanstein Kulz Freifläche"), und jede Erweiterung war entweder
# wirkungslos oder falsch. `ff` als Abkürzung und `bürgerenergie` (eine Rechtsform, keine
# Bauart) sind deshalb draußen, und `park` erst recht: **Parkstein ist ein Ort in der
# Oberpfalz** und käme neunmal fälschlich durch.
#
# Die Anlagenart ist damit im Wesentlichen ein Feld, das in der Projektmaske gepflegt wird. Was
# sich aus PV-, Speicher- und Ladestationsdaten sicher ergeben lässt, leitet der Import ab; ob
# eine 1.364-kWp-Anlage auf einem Hallendach oder auf einer Wiese steht, steht in keiner Spalte.
_FREIFLAECHE = re.compile(
    r"(?<![a-zäöüß])(freifl(ae|ä)che|freiland|solarpark|pv[- ]?park)(?![a-zäöüß])",
    re.IGNORECASE,
)
_EXCEL_NULLTAG = date(1899, 12, 30)  # Excel zählt ab 1900 mit dem bekannten Schaltjahrfehler


class BlattFehlt(FachFehler):
    """Die erwartete Registerkarte ist nicht in der Datei.

    Der häufigste Grund ist eine verwechselte oder umbenannte Datei. Die Meldung nennt deshalb,
    welche Blätter tatsächlich vorhanden sind – das erspart das Rätselraten.
    """

    code = "migration_blatt_fehlt"

    def __init__(self, pfad: Path, erwartet: str, vorhanden: list[str]) -> None:
        super().__init__(
            f"Die Datei '{pfad.name}' hat kein Blatt '{erwartet}'.",
            "Vorhandene Blätter: "
            + (", ".join(f"'{b}'" for b in vorhanden) or "keine")
            + ". Bitte prüfen, ob es die richtige Datei ist.",
        )


@dataclass(frozen=True)
class Befund:
    """Ein Wert, den der Leser nicht sicher deuten konnte.

    ``schwere`` unterscheidet, was Aufmerksamkeit braucht: ``warnung`` heißt, dass eine Angabe
    fehlt oder unlesbar ist; ``hinweis`` heißt, dass bewusst anders verfahren wurde.
    """

    datei: str
    zeile: int
    spalte: str
    wert: str
    meldung: str
    schwere: str = "warnung"

    def als_text(self) -> str:
        return f"{self.datei} {self.spalte}{self.zeile}: {self.meldung} (Inhalt: {self.wert!r})"


@dataclass
class Markerstand:
    """Auswertung einer Kreuz- oder Kalenderwochenzelle."""

    erledigt: bool | None = None
    geplant_kw: str | None = None
    roh: str = ""


@dataclass
class AuftragsZeile:
    """Eine Zeile der Auftragsliste – eine geplante oder gestellte Rechnung."""

    zeile: int
    freitext: str
    kundenteil: str
    kunde: str
    ort: str | None
    rechnungsart: Rechnungsart
    betrag_cent: int
    plan_monat: str | None
    gestellt: bool

    @property
    def ist_projektsumme(self) -> bool:
        """Zeile ohne Rechnungsart: eine Auftragssumme ohne Zahlungsplan (PLAN §9)."""
        return not self.rechnungsart.erkannt


@dataclass
class ProjektZeile:
    """Eine Projektzeile der Teamliste."""

    zeile: int
    kundenteil: str
    kunde: str
    ort: str | None
    pv_kwp: Decimal | None = None
    wr_typ: str | None = None
    speicher_typ: str | None = None
    speicher_kwh: Decimal | None = None
    ladestation: str | None = None
    anlagenart: str | None = None
    auftrag_vom: date | None = None
    ab_wert_cent: int | None = None
    pl_name: str | None = None
    bemerkung: str | None = None
    meilensteine: dict[str, Markerstand] = field(default_factory=dict)
    abschlaege_gestellt: dict[tuple[str, int], bool] = field(default_factory=dict)
    nachkalkulation: dict[str, str] = field(default_factory=dict)
    vorplanung: dict[str, str] = field(default_factory=dict)

    @property
    def status(self) -> str:
        """Projektstatus nach PLAN §9: Abnahme gesetzt heißt abgeschlossen."""
        if self.meilensteine.get("abnahme", Markerstand()).erledigt:
            return "abgeschlossen"
        if any(
            stand.erledigt for typ, stand in self.meilensteine.items() if typ.startswith("montage")
        ):
            return "in_bau"
        return "beauftragt"


@dataclass
class Auftragsliste:
    """Ergebnis des Lesens der Auftragsliste."""

    datei: Path
    zeilen: list[AuftragsZeile] = field(default_factory=list)
    befunde: list[Befund] = field(default_factory=list)

    @property
    def summe_netto_cent(self) -> int:
        return sum(z.betrag_cent for z in self.zeilen)

    @property
    def summe_gestellt_cent(self) -> int:
        return sum(z.betrag_cent for z in self.zeilen if z.gestellt)

    def summe_je_monat(self) -> dict[str, int]:
        """Nettosummen je Planmonat; ``None`` als Schlüssel ``'unterminiert'``."""
        je_monat: dict[str, int] = {}
        for zeile in self.zeilen:
            schluessel = zeile.plan_monat or "unterminiert"
            je_monat[schluessel] = je_monat.get(schluessel, 0) + zeile.betrag_cent
        return dict(sorted(je_monat.items()))


@dataclass
class Teamliste:
    """Ergebnis des Lesens der Teamliste."""

    datei: Path
    zeilen: list[ProjektZeile] = field(default_factory=list)
    befunde: list[Befund] = field(default_factory=list)

    @property
    def summe_ab_wert_cent(self) -> int:
        return sum(z.ab_wert_cent or 0 for z in self.zeilen)

    @property
    def summe_pv_kwp(self) -> Decimal:
        return sum((z.pv_kwp or Decimal(0) for z in self.zeilen), Decimal(0))

    def anzahl_je_status(self) -> dict[str, int]:
        je_status: dict[str, int] = {}
        for zeile in self.zeilen:
            je_status[zeile.status] = je_status.get(zeile.status, 0) + 1
        return dict(sorted(je_status.items()))


def _zellen(zeile: tuple[Any, ...]) -> dict[str, Any]:
    """Belegte Zellen einer Zeile als ``{Spaltenbuchstabe: Wert}``.

    Im ``read_only``-Modus liefert openpyxl für unbelegte Stellen ``EmptyCell``-Objekte, die
    weder ``column_letter`` noch ``row`` kennen. Sie werden hier weggefiltert.
    """
    belegt: dict[str, Any] = {}
    for zelle in zeile:
        buchstabe = getattr(zelle, "column_letter", None)
        if buchstabe is not None and zelle.value is not None:
            belegt[buchstabe] = zelle.value
    return belegt


def _text(wert: Any) -> str:
    """Zellinhalt als getrimmter Text; ``None`` wird zum Leerstring."""
    if wert is None:
        return ""
    if isinstance(wert, str):
        return wert.strip()
    return str(wert).strip()


def _ist_fehlerwert(text: str) -> bool:
    """Excel-Fehlerwerte wie ``#VALUE!`` oder ``#REF!``."""
    return text.startswith("#") and text.endswith("!")


def _zahl(text: str) -> Decimal | None:
    """Dezimalzahl aus einem Zellinhalt, oder ``None`` wenn es keine ist."""
    if not text:
        return None
    try:
        return Decimal(text.replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def marker_lesen(wert: Any) -> Markerstand:
    """Deutet eine Zelle des Termin- oder Statusblocks.

    Die Blöcke sind über Jahre gemischt gepflegt worden: meist ``x`` und ``-``, daneben ``o``
    für offen und in einigen Zeilen eine Kalenderwoche in der Form ``28/22``. Was keiner Form
    entspricht, bleibt ``erledigt=None`` und wird von der aufrufenden Stelle zum Befund gemacht.
    """
    roh = _text(wert)
    if not roh:
        return Markerstand(roh="")
    if roh in MARKER_JA:
        return Markerstand(erledigt=True, roh=roh)
    if roh in MARKER_NEIN:
        return Markerstand(erledigt=False, roh=roh)
    if roh in MARKER_OFFEN:
        # 'o' steht für offen: der Schritt ist benannt, aber nicht erledigt.
        return Markerstand(erledigt=False, roh=roh)
    treffer = _KALENDERWOCHE.match(roh)
    if treffer:
        woche, jahr = treffer.groups()
        return Markerstand(geplant_kw=f"{int(woche):02d}/{jahr[-2:]}", roh=roh)
    # Mehrfachkreuze wie 'x, x' sind ein Kreuz mit Tippfehler – erledigt, aber auffällig.
    if set(roh.replace(",", " ").split()) <= {"x", "X"}:
        return Markerstand(erledigt=True, roh=roh)
    return Markerstand(roh=roh)


def excel_datum(wert: Any) -> date | None:
    """Excel-Datumszahl oder echtes Datum in ein ``date`` wandeln.

    Die Auftragsspalte der Teamliste enthält Seriennummern (44123), weil die Datei alt ist.
    openpyxl wandelt formatierte Zellen selbst um; unformatierte kommen als Zahl an.
    """
    if wert is None:
        return None
    if isinstance(wert, date):
        return wert
    zahl = _zahl(_text(wert))
    if zahl is None:
        return None
    tage = int(zahl)
    # Alles unter 1000 ist keine plausible Datumszahl (1902), sondern ein Tippfehler.
    if tage < 1000 or tage > 80000:
        return None
    return _EXCEL_NULLTAG + timedelta(days=tage)


def auftragsliste_lesen(pfad: Path) -> Auftragsliste:
    """Liest die Auftragsliste (``Offene_Auftraege``, Blatt ``Et-Einnahmen``).

    Gelesen werden die Formeln, nicht die zwischengespeicherten Ergebnisse: die Monatsspalten
    tragen nur ``=WENN(<Marker>="x";$B;"")``, den Planmonat verrät allein die Markerspalte.
    """
    ergebnis = Auftragsliste(datei=pfad)
    name = pfad.stem
    arbeitsmappe = load_workbook(pfad, read_only=True, data_only=False)
    try:
        if BLATT_AUFTRAEGE not in arbeitsmappe.sheetnames:
            raise BlattFehlt(pfad, BLATT_AUFTRAEGE, arbeitsmappe.sheetnames)
        blatt = arbeitsmappe[BLATT_AUFTRAEGE]
        # Die Zeilennummer kommt aus dem Zähler, nicht aus der Zelle: im read_only-Modus haben
        # unbelegte Zellen kein .row. iter_rows liefert die Zeilen lückenlos in Reihenfolge.
        for nummer, zeile in enumerate(
            blatt.iter_rows(min_row=ERSTE_DATENZEILE), start=ERSTE_DATENZEILE
        ):
            zellen = _zellen(zeile)
            if not zellen:
                continue
            freitext = _text(zellen.get(SPALTE_TEXT))
            if not freitext:
                continue

            betrag_text = _text(zellen.get(SPALTE_BETRAG))
            betrag = _zahl(betrag_text)
            if betrag is None:
                ergebnis.befunde.append(
                    Befund(
                        name,
                        nummer,
                        SPALTE_BETRAG,
                        betrag_text,
                        "Kein lesbarer Nettobetrag – Zeile wird nicht übernommen",
                    )
                )
                continue

            kundenteil, art = rechnungsart_lesen(freitext)
            kunde, ort = kunde_und_ort(kundenteil)
            plan_monat, monatsbefund = _plan_monat(zellen, name, nummer)
            if monatsbefund:
                ergebnis.befunde.append(monatsbefund)

            ergebnis.zeilen.append(
                AuftragsZeile(
                    zeile=nummer,
                    freitext=freitext,
                    kundenteil=kundenteil,
                    kunde=kunde,
                    ort=ort,
                    rechnungsart=art,
                    betrag_cent=euro_nach_cent(betrag),
                    plan_monat=plan_monat,
                    gestellt=_text(zellen.get(SPALTE_GESTELLT)) in MARKER_JA,
                )
            )
            if not art.erkannt:
                ergebnis.befunde.append(
                    Befund(
                        name,
                        nummer,
                        SPALTE_TEXT,
                        freitext,
                        "Keine Rechnungsart erkennbar – wird als Auftragssumme ohne "
                        "Zahlungsplan übernommen",
                        schwere="hinweis",
                    )
                )
            elif art.gewerk is None:
                ergebnis.befunde.append(
                    Befund(
                        name,
                        nummer,
                        SPALTE_TEXT,
                        freitext,
                        "Gewerk nicht aus dem Text erkennbar – wird aus den Anlagendaten "
                        "des Projekts abgeleitet",
                        schwere="hinweis",
                    )
                )
    finally:
        arbeitsmappe.close()
    return ergebnis


def _plan_monat(
    zellen: dict[str, Any], datei: str, nummer: int, jahr: int = 2026
) -> tuple[str | None, Befund | None]:
    """Planmonat aus den elf Markerspalten.

    Das Jahr ist nicht in der Datei vermerkt; 2026 ist von Sven bestätigt (siehe
    docs/OFFENE-PUNKTE.md Nr. 8). Mehrere Kreuze in einer Zeile kommen im Bestand nicht vor –
    falls doch, gewinnt der erste und die Zeile wird zum Befund.
    """
    gesetzt = [
        i
        for i, spalte in enumerate(MARKERSPALTEN, start=1)
        if _text(zellen.get(spalte)) in MARKER_JA
    ]
    if not gesetzt:
        return None, None
    monat = f"{jahr}-{gesetzt[0]:02d}"
    if len(gesetzt) > 1:
        monate = ", ".join(f"{jahr}-{m:02d}" for m in gesetzt)
        return monat, Befund(
            datei,
            nummer,
            "".join(MARKERSPALTEN[m - 1] for m in gesetzt),
            monate,
            f"Mehrere Planmonate angekreuzt – übernommen wird {monat}",
        )
    return monat, None


def teamliste_lesen(pfad: Path) -> Teamliste:
    """Liest die Teamliste (``Teambesprechung_NEU``, Blatt ``Übersicht Projekte``)."""
    ergebnis = Teamliste(datei=pfad)
    name = pfad.stem
    # data_only=True: hier interessieren die Werte, nicht die Formeln. Fehlerwerte kommen als
    # Text ('#VALUE!') an und werden als Befund gemeldet.
    arbeitsmappe = load_workbook(pfad, read_only=True, data_only=True)
    try:
        if BLATT_PROJEKTE not in arbeitsmappe.sheetnames:
            raise BlattFehlt(pfad, BLATT_PROJEKTE, arbeitsmappe.sheetnames)
        blatt = arbeitsmappe[BLATT_PROJEKTE]
        for nummer, zeile in enumerate(
            blatt.iter_rows(min_row=ERSTE_DATENZEILE), start=ERSTE_DATENZEILE
        ):
            zellen = _zellen(zeile)
            kundentext = _text(zellen.get(SPALTE_KUNDE))
            if not kundentext:
                continue
            kunde, ort = kunde_und_ort(kundentext)
            projekt = ProjektZeile(zeile=nummer, kundenteil=kundentext, kunde=kunde, ort=ort)
            _stammdaten_lesen(projekt, zellen, ergebnis.befunde, name)
            _meilensteine_lesen(projekt, zellen, ergebnis.befunde, name)
            _nebenspalten_lesen(projekt, zellen, ergebnis.befunde, name)
            ergebnis.zeilen.append(projekt)
    finally:
        arbeitsmappe.close()
    return ergebnis


def _stammdaten_lesen(
    projekt: ProjektZeile, zellen: dict[str, Any], befunde: list[Befund], datei: str
) -> None:
    """Anlagendaten, Auftragsdatum, AB-Wert und Projektleiter."""
    projekt.pv_kwp = _zahl(_text(zellen.get(SPALTE_PV_KWP)))
    projekt.wr_typ = _freitext_oder_none(zellen.get(SPALTE_WR))
    projekt.ladestation = _freitext_oder_none(zellen.get(SPALTE_LADESTATION))

    # Die Speicherspalte führt eine Produktbezeichnung ('2x BYD HVM 22.1'), keine Zahl. Die
    # Kapazität steckt darin und wird herausgelesen; der Text bleibt als Gerätebezeichnung.
    speicher = _freitext_oder_none(zellen.get(SPALTE_SPEICHER))
    if speicher:
        projekt.speicher_typ = speicher
        projekt.speicher_kwh = _kapazitaet_aus_text(speicher)
        if projekt.speicher_kwh is None:
            befunde.append(
                Befund(
                    datei,
                    projekt.zeile,
                    SPALTE_SPEICHER,
                    speicher,
                    "Keine Speicherkapazität aus der Bezeichnung lesbar – "
                    "Bezeichnung wird übernommen, kWh bleibt leer",
                    schwere="hinweis",
                )
            )
    # Spalte F ('Storage') ist in fünf Zeilen gefüllt und benennt dasselbe Gerät anders. Sie
    # ergänzt die Bezeichnung, überschreibt sie aber nicht.
    storage = _freitext_oder_none(zellen.get(SPALTE_STORAGE))
    if storage:
        projekt.speicher_typ = (
            f"{projekt.speicher_typ}; {storage}" if projekt.speicher_typ else storage
        )
        projekt.speicher_kwh = projekt.speicher_kwh or _kapazitaet_aus_text(storage)

    projekt.anlagenart, per_stichwort = anlagenart_ableiten(
        projekt.pv_kwp,
        projekt.speicher_typ,
        projekt.ladestation,
        f"{projekt.kundenteil} {projekt.ort or ''}",
    )
    if per_stichwort:
        befunde.append(
            Befund(
                datei,
                projekt.zeile,
                SPALTE_KUNDE,
                projekt.kundenteil,
                "Anlagenart 'Freifläche' aus dem Namen erschlossen – kein Feld der Teamliste "
                "sagt das. Bitte in der Projektmaske bestätigen",
                schwere="hinweis",
            )
        )

    datum_roh = _text(zellen.get(SPALTE_AUFTRAG_VOM))
    projekt.auftrag_vom = excel_datum(zellen.get(SPALTE_AUFTRAG_VOM))
    if datum_roh and projekt.auftrag_vom is None and datum_roh not in MARKER_NEIN:
        befunde.append(
            Befund(
                datei,
                projekt.zeile,
                SPALTE_AUFTRAG_VOM,
                datum_roh,
                "Kein lesbares Auftragsdatum – die Projektnummer wird dem laufenden Jahr "
                "zugeordnet",
            )
        )

    ab_roh = _text(zellen.get(SPALTE_AB_WERT))
    ab_wert = _zahl(ab_roh)
    if ab_wert is not None:
        projekt.ab_wert_cent = euro_nach_cent(ab_wert)
    elif ab_roh and ab_roh not in MARKER_NEIN:
        # '22.604.28 €' und '4.999.59' sind Tippfehler mit zwei Trennzeichen. Der Betrag wird
        # nicht geraten: bei Geld ist eine Lücke besser als eine erfundene Zahl (PLAN §6).
        befunde.append(
            Befund(
                datei,
                projekt.zeile,
                SPALTE_AB_WERT,
                ab_roh,
                "Kein lesbarer Auftragswert – bleibt leer und ist in der Maske nachzutragen",
            )
        )

    pl = _freitext_oder_none(zellen.get(SPALTE_PL))
    projekt.pl_name = pl
    projekt.bemerkung = _freitext_oder_none(zellen.get(SPALTE_BEMERKUNG))


def anlagenart_ableiten(
    pv_kwp: Decimal | None,
    speicher: str | None,
    ladestation: str | None,
    freitext: str,
) -> tuple[str | None, bool]:
    """Anlagenart für Liste und Filter (design/Projektliste.dc.html).

    Rückgabe ist ``(art, per_stichwort)``. ``per_stichwort`` sagt, ob die Freifläche über den
    Klartext erkannt wurde – dann ist die Angabe geraten und gehört ins Protokoll.

    Ohne PV, Speicher und Ladestation bleibt die Art leer: ``sonstig`` wäre eine Aussage, und
    hier ist nichts bekannt.
    """
    hat_pv = pv_kwp is not None
    hat_speicher = speicher is not None
    per_stichwort = bool(_FREIFLAECHE.search(freitext))

    if hat_pv and per_stichwort:
        return "freiflaeche", True
    if hat_pv and hat_speicher:
        return "aufdach_speicher", False
    if hat_pv:
        return "aufdach", False
    if hat_speicher:
        return "speicher", False
    if ladestation is not None:
        return "ladestation", False
    return None, False


def _freitext_oder_none(wert: Any) -> str | None:
    """Freitextzelle: leere Zellen, Striche und Excel-Fehlerwerte gelten als ohne Angabe."""
    text = _text(wert)
    if not text or text in MARKER_NEIN or _ist_fehlerwert(text):
        return None
    return text


def _kapazitaet_aus_text(text: str) -> Decimal | None:
    """Speicherkapazität aus einer Produktbezeichnung.

    '2x BYD HVM 22.1' sind zwei Geräte mit je 22,1 kWh, also 44,2 kWh. 'BYD HVM 11.0' ist eines
    mit 11 kWh. Gelesen wird die letzte Dezimalzahl, weil die Bezeichnung mit ihr endet;
    ein vorangestelltes 'nx' vervielfacht sie.
    """
    zahlen = re.findall(r"\d+[.,]\d+", text)
    if not zahlen:
        return None
    kapazitaet = _zahl(zahlen[-1])
    if kapazitaet is None:
        return None
    vielfach = re.match(r"\s*(\d+)\s*[x×]", text, re.IGNORECASE)
    if vielfach:
        kapazitaet *= int(vielfach.group(1))
    return kapazitaet


def _meilensteine_lesen(
    projekt: ProjektZeile, zellen: dict[str, Any], befunde: list[Befund], datei: str
) -> None:
    """Termin- und Statusblock in Meilensteine übersetzen."""
    for spalte, typ in {**TERMINSPALTEN, **STATUSSPALTEN}.items():
        stand = marker_lesen(zellen.get(spalte))
        if not stand.roh:
            continue
        projekt.meilensteine[typ] = stand
        if stand.erledigt is None and stand.geplant_kw is None:
            befunde.append(
                Befund(
                    datei,
                    projekt.zeile,
                    spalte,
                    stand.roh,
                    f"Zelle im Termin- oder Statusblock ({typ}) ist weder Kreuz noch "
                    "Kalenderwoche – Meilenstein wird ohne Stand angelegt",
                )
            )


def _nebenspalten_lesen(
    projekt: ProjektZeile, zellen: dict[str, Any], befunde: list[Befund], datei: str
) -> None:
    """Abschlagskreuze, Nachkalkulations-Altwerte und Vorplanungsblock."""
    for spalte, (gewerk, nummer) in ABSCHLAGSSPALTEN.items():
        stand = marker_lesen(zellen.get(spalte))
        if stand.erledigt is not None:
            projekt.abschlaege_gestellt[(gewerk, nummer)] = stand.erledigt

    for spalte, bezeichnung in NACHKALKULATION.items():
        wert = _freitext_oder_none(zellen.get(spalte))
        if wert:
            projekt.nachkalkulation[bezeichnung] = wert

    for spalte, bezeichnung in VORPLANUNGSSPALTEN.items():
        wert = _freitext_oder_none(zellen.get(spalte))
        if not wert:
            continue
        projekt.vorplanung[bezeichnung] = wert
        befunde.append(
            Befund(
                datei,
                projekt.zeile,
                spalte,
                wert,
                f"Vorplanungsspalte '{bezeichnung}' wird als Notiz am Projekt geführt, "
                "nicht als Meilenstein",
                schwere="hinweis",
            )
        )

    module = _text(zellen.get(SPALTE_MODULE_RESERVIERT))
    if module and _ist_fehlerwert(module):
        befunde.append(
            Befund(
                datei,
                projekt.zeile,
                SPALTE_MODULE_RESERVIERT,
                module,
                "Fehlerwert in 'Module reserviert' – die Spalte wird ohnehin nicht übernommen",
                schwere="hinweis",
            )
        )
