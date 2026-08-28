"""Kalkulationsblatt einlesen und die Vorlage dafür erzeugen (PLAN §8, EXPORT-Tab).

Sven kalkuliert seine Projekte in Excel. Damit der Leitstand die Sollwerte übernehmen kann, ohne
sich in eine über Jahre gewachsene Tabelle einzumischen, bekommt jedes Kalkulationsblatt ein
zusätzliches Blatt ``EXPORT`` mit **benannten Zellen**. Gelesen wird über diese Namen, nie über
Zellkoordinaten: so darf im Blatt jederzeit eine Zeile eingefügt werden, ohne dass der Import
danebengreift. Die Vorlage dafür erzeugt dieses Modul (PLAN §8: „Claude Code erzeugt die
Vorlagendatei"); Sven verknüpft die Zellen einmalig mit seiner eigenen Kalkulation.

Arbeitsteilung bei Fehlern:

* **Baufehler** – kein Blatt ``EXPORT``, fehlende Namen – sind ein :class:`KalkulationsblattFehler`.
  Wer eine einzelne Datei ausgewählt hat, bekommt so sofort gesagt, was daran fehlt. Der
  nächtliche Ordnerlauf fängt ihn je Datei ab und macht einen Befund daraus, damit eine kaputte
  Datei nicht den ganzen Lauf verhindert.
* **Inhaltsfehler** – unlesbare Menge, unbekanntes Gewerk, fehlende Projektnummer – ergeben einen
  :class:`~app.importe.befunde.Befund`, und der Rest der Datei wird trotzdem gelesen.

Beim Übernehmen gilt eine Regel, die zählt (siehe :func:`uebernehmen`): **eine bestätigte
Ist-Menge wird nie überschrieben.** Das Kalkulationsblatt ist die Wahrheit für das Soll, die
Maske „Mengen-Ist bestätigen" die Wahrheit für das Ist.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_to_tuple, get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.fehler import FachFehler
from app.geld import euro_nach_cent
from app.importe.befunde import Befund
from app.importe.werte import ist_fehlerwert, text, zahl
from app.modelle import Projekt, SollKalkulation, Stuecklistenposition
from app.modelle.kalkulation import STUECKLISTE_QUELLEN
from app.zeit import jetzt_utc

BLATT = "EXPORT"

NAME_PROJEKT_NR = "exp_projekt_nr"
NAME_MATERIAL_SOLL = "exp_material_soll"
NAME_DL_SOLL = "exp_dl_soll"
NAME_STUNDEN_SOLL = "exp_stunden_soll"
NAME_MARGE_SOLL = "exp_marge_soll"
NAME_POSITIONEN_START = "exp_positionen_start"

PFLICHTNAMEN: tuple[str, ...] = (
    NAME_PROJEKT_NR,
    NAME_MATERIAL_SOLL,
    NAME_DL_SOLL,
    NAME_STUNDEN_SOLL,
    NAME_MARGE_SOLL,
    NAME_POSITIONEN_START,
)

# Spalten der Positionstabelle, ab ``exp_positionen_start`` nach rechts (PLAN §8).
SPALTEN_POSITIONEN: tuple[str, ...] = (
    "artikel_nr",
    "bezeichnung",
    "menge",
    "ep_ek",
    "quelle",
    "gewerk",
)

# Gewerke, die eine Stücklistenposition tragen darf. 'service' und 'nachtrag' aus GEWERKE
# gehören zum Zahlungsplan, nicht zur Kalkulation (siehe Prüfung am Modell).
GEWERKE_KALKULATION: tuple[str, ...] = ("pv", "speicher", "ls")

# Nach so vielen leeren Zeilen hintereinander endet die Positionstabelle. Eine einzelne
# Leerzeile als Gliederung soll die Liste nicht abschneiden.
LEERZEILEN_ENDE = 5

# Obergrenze, damit ein Blatt mit einer Formel über 100.000 Zeilen den Lauf nicht aufhält.
MAX_POSITIONEN = 5000

# Prozent in Promille: marge_soll steht als Integer in Promille in der Datenbank, damit die
# Ampel ohne Gleitkomma vergleicht (siehe app/modelle/kalkulation.py).
PROMILLE_JE_PROZENT = 10

# Dateiname beginnt mit der Projektnummer (PLAN §8): '26001_Mustermann.xlsx'.
# Kein \b hinter der Nummer: zwischen Ziffer und Unterstrich steht keine Wortgrenze, und der
# Unterstrich ist genau die Schreibweise aus PLAN §8. Stattdessen: keine weitere Ziffer danach,
# damit aus '260011_...' nicht die Nummer 26001 gelesen wird.
_DATEINAME = re.compile(r"^(\d{4,8})(?!\d)")


class KalkulationsblattFehler(FachFehler):
    """Basis für alles, was am Aufbau der Datei liegt."""

    code = "kalkulationsblatt_fehler"


class ExportBlattFehlt(KalkulationsblattFehler):
    code = "kalkulationsblatt_export_fehlt"

    def __init__(self, pfad: Path, vorhanden: list[str]) -> None:
        super().__init__(
            f"Die Datei '{pfad.name}' hat kein Blatt '{BLATT}'.",
            "Vorhandene Blätter: "
            + (", ".join(f"'{b}'" for b in vorhanden) or "keine")
            + f". Die Vorlage mit dem Blatt '{BLATT}' liegt unter vorlagen/ und lässt sich mit "
            "'ip3-leitstand kalkulationsblatt-vorlage' neu erzeugen.",
        )


class NamenFehlen(KalkulationsblattFehler):
    """Das Blatt ist da, aber die benannten Zellen fehlen.

    Der häufigste Grund: das Blatt wurde aus der Vorlage herauskopiert, dabei gehen benannte
    Zellen verloren. Die Meldung nennt alle fehlenden auf einmal – sonst wird das ein
    Wechselspiel aus Probieren und Nachbessern.
    """

    code = "kalkulationsblatt_namen_fehlen"

    def __init__(self, pfad: Path, fehlend: list[str]) -> None:
        super().__init__(
            f"Im Blatt '{BLATT}' der Datei '{pfad.name}' fehlen benannte Zellen: "
            + ", ".join(fehlend)
            + ".",
            "In Excel unter Formeln → Namensmanager anlegen und je auf die Zelle mit dem Wert "
            "zeigen lassen. Am einfachsten ist es, das Blatt aus vorlagen/"
            "Kalkulationsblatt-Vorlage.xlsx zu übernehmen und dort die Zellen mit der eigenen "
            "Kalkulation zu verknüpfen.",
        )


class WerteFehlen(KalkulationsblattFehler):
    """Die Datei enthält Formeln, aber keine berechneten Werte.

    Excel legt die Ergebnisse beim Speichern ab. Eine Datei, die ein Programm erzeugt und nie
    ein Mensch geöffnet hat, trägt nur die Formeln – der Import läse dann überall nichts und
    würde stumm ein leeres Kalkulationsblatt melden.
    """

    code = "kalkulationsblatt_werte_fehlen"

    def __init__(self, pfad: Path, zellen: list[str]) -> None:
        super().__init__(
            f"In der Datei '{pfad.name}' stehen Formeln ohne berechnetes Ergebnis "
            f"({', '.join(zellen)}).",
            "Die Datei einmal in Excel öffnen, speichern und schließen. Danach steht das "
            "Ergebnis in der Datei und der Leitstand kann es lesen.",
        )


@dataclass
class Kalkulationsposition:
    """Eine Zeile der Stückliste aus dem EXPORT-Blatt."""

    zeile: int
    bezeichnung: str
    menge: Decimal
    quelle: str
    artikel_nr: str | None = None
    ek_preis_cent: int | None = None
    gewerk: str | None = None

    @property
    def schluessel(self) -> str:
        """Vergleichsschlüssel für den Abgleich mit der gespeicherten Stückliste.

        Artikelnummer, wenn es eine gibt – sonst die Bezeichnung. Ohne Artikelnummer ist die
        Bezeichnung das Einzige, woran eine Position wiederzuerkennen ist.
        """
        if self.artikel_nr:
            return f"a:{self.artikel_nr.casefold()}"
        return f"b:{' '.join(self.bezeichnung.casefold().split())}"


@dataclass
class Kalkulationsblatt:
    """Der gelesene Inhalt eines EXPORT-Blattes. Nichts davon ist geschrieben."""

    datei: Path
    projekt_nr: int | None = None
    material_soll_cent: int | None = None
    dl_soll_cent: int | None = None
    stunden_soll: Decimal | None = None
    marge_soll_promille: int | None = None
    positionen: list[Kalkulationsposition] = field(default_factory=list)
    befunde: list[Befund] = field(default_factory=list)

    @property
    def soll_gesamt_cent(self) -> int | None:
        """Material plus Dienstleistung, oder ``None``, wenn beide fehlen."""
        if self.material_soll_cent is None and self.dl_soll_cent is None:
            return None
        return (self.material_soll_cent or 0) + (self.dl_soll_cent or 0)

    @property
    def lagerpositionen(self) -> list[Kalkulationsposition]:
        """Positionen, die aus dem Lager kommen und damit zu bewerten sind (PLAN §6.5)."""
        return [p for p in self.positionen if p.quelle == "lager"]


# ---------------------------------------------------------------------------
# Lesen
# ---------------------------------------------------------------------------


def blatt_lesen(pfad: Path) -> Kalkulationsblatt:
    """Liest das Blatt ``EXPORT`` einer Kalkulationsdatei. Schreibt nichts."""
    ergebnis = Kalkulationsblatt(datei=pfad)
    werte = load_workbook(pfad, data_only=True)
    formeln = load_workbook(pfad, data_only=False)
    try:
        if BLATT not in werte.sheetnames:
            raise ExportBlattFehlt(pfad, list(werte.sheetnames))
        blatt = werte[BLATT]
        ziele = _namen_aufloesen(werte, pfad)
        _auf_fehlende_ergebnisse_pruefen(werte, formeln, ziele, pfad)

        ergebnis.projekt_nr = _projekt_nr_lesen(blatt, ziele, ergebnis)
        ergebnis.material_soll_cent = _geld_lesen(
            blatt, ziele, NAME_MATERIAL_SOLL, "Materialkosten Soll", ergebnis
        )
        ergebnis.dl_soll_cent = _geld_lesen(
            blatt, ziele, NAME_DL_SOLL, "Dienstleistung Soll", ergebnis
        )
        ergebnis.stunden_soll = _stunden_lesen(blatt, ziele, ergebnis)
        ergebnis.marge_soll_promille = _marge_lesen(blatt, ziele, ergebnis)
        ergebnis.positionen = _positionen_lesen(werte, ziele, ergebnis)
    finally:
        werte.close()
        formeln.close()
    return ergebnis


def _namen_aufloesen(mappe: Workbook, pfad: Path) -> dict[str, tuple[str, str]]:
    """Benannte Zellen zu ``{Name: (Blattname, Koordinate)}``.

    Gesucht wird erst unter den Namen der Arbeitsmappe, dann unter den blattlokalen – Excel legt
    sie je nach Bedienweg an der einen oder der anderen Stelle ab.
    """
    ziele: dict[str, tuple[str, str]] = {}
    fehlend: list[str] = []
    lokal = mappe[BLATT].defined_names if BLATT in mappe.sheetnames else {}
    for name in PFLICHTNAMEN:
        eintrag = mappe.defined_names.get(name) or lokal.get(name)
        stellen = list(eintrag.destinations) if eintrag is not None else []
        if not stellen:
            fehlend.append(name)
            continue
        ziele[name] = stellen[0]
    if fehlend:
        raise NamenFehlen(pfad, fehlend)
    return ziele


def _auf_fehlende_ergebnisse_pruefen(
    werte: Workbook, formeln: Workbook, ziele: dict[str, tuple[str, str]], pfad: Path
) -> None:
    """Formel ohne gespeichertes Ergebnis erkennen, bevor der Import stumm nichts findet."""
    betroffen: list[str] = []
    for name, (blattname, koordinate) in ziele.items():
        if name == NAME_POSITIONEN_START:
            continue
        if werte[blattname][koordinate].value is not None:
            continue
        roh = formeln[blattname][koordinate].value
        if isinstance(roh, str) and roh.startswith("="):
            betroffen.append(f"{name} ({blattname}!{koordinate.replace('$', '')})")
    if betroffen:
        raise WerteFehlen(pfad, betroffen)


def _zelle(blatt: Any, ziele: dict[str, tuple[str, str]], name: str) -> Any:
    return blatt.parent[ziele[name][0]][ziele[name][1]].value


def _befund(ergebnis: Kalkulationsblatt, name: str, wert: Any, meldung: str) -> None:
    ergebnis.befunde.append(
        Befund(
            datei=ergebnis.datei.name,
            zeile=0,
            spalte=name,
            wert=text(wert),
            meldung=meldung,
        )
    )


def _projekt_nr_lesen(
    blatt: Any, ziele: dict[str, tuple[str, str]], ergebnis: Kalkulationsblatt
) -> int | None:
    roh = _zelle(blatt, ziele, NAME_PROJEKT_NR)
    inhalt = text(roh)
    wert = zahl(inhalt)
    if wert is None or wert != wert.to_integral_value() or wert <= 0:
        _befund(
            ergebnis,
            NAME_PROJEKT_NR,
            roh,
            "Keine lesbare Projektnummer – die Datei wird nicht übernommen",
        )
        return None
    return int(wert)


def _geld_lesen(
    blatt: Any,
    ziele: dict[str, tuple[str, str]],
    name: str,
    bezeichnung: str,
    ergebnis: Kalkulationsblatt,
) -> int | None:
    roh = _zelle(blatt, ziele, name)
    inhalt = text(roh)
    if not inhalt:
        return None
    if ist_fehlerwert(inhalt):
        _befund(ergebnis, name, roh, f"{bezeichnung}: Excel-Fehlerwert in der Zelle")
        return None
    wert = zahl(inhalt)
    if wert is None:
        _befund(ergebnis, name, roh, f"{bezeichnung}: kein lesbarer Betrag")
        return None
    if wert < 0:
        _befund(ergebnis, name, roh, f"{bezeichnung}: negativer Betrag, wird nicht übernommen")
        return None
    return euro_nach_cent(wert)


def _stunden_lesen(
    blatt: Any, ziele: dict[str, tuple[str, str]], ergebnis: Kalkulationsblatt
) -> Decimal | None:
    roh = _zelle(blatt, ziele, NAME_STUNDEN_SOLL)
    inhalt = text(roh)
    if not inhalt:
        return None
    wert = zahl(inhalt)
    if wert is None or wert < 0:
        _befund(ergebnis, NAME_STUNDEN_SOLL, roh, "Keine lesbare Stundenzahl")
        return None
    return wert.quantize(Decimal("0.01"))


def _marge_lesen(
    blatt: Any, ziele: dict[str, tuple[str, str]], ergebnis: Kalkulationsblatt
) -> int | None:
    """Sollmarge in Promille.

    Excel speichert eine prozentformatierte Zelle als Bruchteil (18 % werden zu 0,18), eine
    einfach getippte Zahl dagegen als 18. Beides kommt vor, also werden beide gedeutet: ein Wert
    **unter 1** gilt als Bruchteil, ein Wert **ab 1** als Prozent. Die Grenze liegt damit bei
    einer Sollmarge von 1 % – darunter rechnet niemand, und 100 % wäre keine Marge, sondern ein
    Projekt ohne Kosten.
    """
    roh = _zelle(blatt, ziele, NAME_MARGE_SOLL)
    inhalt = text(roh)
    if not inhalt:
        return None
    wert = zahl(inhalt.rstrip("% ").strip())
    if wert is None:
        _befund(ergebnis, NAME_MARGE_SOLL, roh, "Keine lesbare Sollmarge")
        return None
    prozent = wert * 100 if 0 < wert < 1 else wert
    if not -100 <= prozent <= 100:
        _befund(
            ergebnis,
            NAME_MARGE_SOLL,
            roh,
            "Sollmarge außerhalb von −100 % bis 100 % – bitte die Zelle prüfen",
        )
        return None
    return int((prozent * PROMILLE_JE_PROZENT).to_integral_value())


def _positionen_lesen(
    mappe: Workbook, ziele: dict[str, tuple[str, str]], ergebnis: Kalkulationsblatt
) -> list[Kalkulationsposition]:
    blattname, koordinate = ziele[NAME_POSITIONEN_START]
    blatt = mappe[blattname]
    erste_zeile, erste_spalte = coordinate_to_tuple(koordinate.replace("$", ""))

    positionen: list[Kalkulationsposition] = []
    leer_am_stueck = 0
    for versatz in range(MAX_POSITIONEN):
        nummer = erste_zeile + versatz
        felder = {
            spalte: blatt.cell(row=nummer, column=erste_spalte + i).value
            for i, spalte in enumerate(SPALTEN_POSITIONEN)
        }
        if all(not text(w) for w in felder.values()):
            leer_am_stueck += 1
            if leer_am_stueck >= LEERZEILEN_ENDE:
                break
            continue
        leer_am_stueck = 0
        position = _position_deuten(felder, nummer, ergebnis)
        if position is not None:
            positionen.append(position)
    return positionen


def _position_deuten(
    felder: dict[str, Any], nummer: int, ergebnis: Kalkulationsblatt
) -> Kalkulationsposition | None:
    """Eine Zeile der Positionstabelle. ``None`` heißt: nicht übernommen, Befund steht schon."""

    def zeilenbefund(spalte: str, meldung: str) -> None:
        ergebnis.befunde.append(
            Befund(
                datei=ergebnis.datei.name,
                zeile=nummer,
                spalte=spalte,
                wert=text(felder.get(spalte)),
                meldung=meldung,
            )
        )

    bezeichnung = text(felder["bezeichnung"])
    if not bezeichnung:
        zeilenbefund("bezeichnung", "Position ohne Bezeichnung – Zeile wird nicht übernommen")
        return None

    menge = zahl(text(felder["menge"]))
    if menge is None:
        zeilenbefund("menge", "Keine lesbare Menge – Zeile wird nicht übernommen")
        return None

    quelle = text(felder["quelle"]).casefold()
    if quelle not in STUECKLISTE_QUELLEN:
        zeilenbefund(
            "quelle",
            "Quelle muss "
            + " oder ".join(f"'{q}'" for q in STUECKLISTE_QUELLEN)
            + " sein – Zeile wird nicht übernommen, weil sonst offen bliebe, ob das Material "
            "über DATEV oder über die Lagerbewertung ins Projekt kommt",
        )
        return None

    gewerk: str | None = text(felder["gewerk"]).casefold() or None
    if gewerk is not None and gewerk not in GEWERKE_KALKULATION:
        zeilenbefund(
            "gewerk",
            "Unbekanntes Gewerk – die Position wird ohne Gewerk übernommen. Zulässig: "
            + ", ".join(GEWERKE_KALKULATION),
        )
        gewerk = None

    ek_preis_cent: int | None = None
    ek_roh = text(felder["ep_ek"])
    if ek_roh:
        ek = zahl(ek_roh)
        if ek is None:
            zeilenbefund("ep_ek", "Kein lesbarer Einkaufspreis – die Position bleibt unbewertet")
        elif ek < 0:
            zeilenbefund("ep_ek", "Negativer Einkaufspreis – die Position bleibt unbewertet")
        else:
            ek_preis_cent = euro_nach_cent(ek)
    elif quelle == "lager":
        zeilenbefund(
            "ep_ek",
            "Lagerposition ohne Einkaufspreis – sie kann nicht bewertet werden und fehlt "
            "damit im Ist (PLAN §6.5)",
        )

    return Kalkulationsposition(
        zeile=nummer,
        artikel_nr=text(felder["artikel_nr"]) or None,
        bezeichnung=bezeichnung,
        menge=menge.quantize(Decimal("0.001")),
        ek_preis_cent=ek_preis_cent,
        quelle=quelle,
        gewerk=gewerk,
    )


# ---------------------------------------------------------------------------
# Ordner scannen
# ---------------------------------------------------------------------------


@dataclass
class Kalkulationsdatei:
    projekt_nr: int
    pfad: Path


def ordner_scannen(ordner: Path) -> tuple[list[Kalkulationsdatei], list[Befund]]:
    """Kalkulationsblätter in ``03_Kalkulation`` finden (PLAN §8).

    Zugeordnet wird über die führende Projektnummer im Dateinamen. Liegen mehrere Dateien zu
    einem Projekt, gewinnt die zuletzt geänderte; die übrigen stehen als Hinweis im Protokoll –
    stillschweigend die falsche zu nehmen wäre schlimmer als die Nachfrage.
    """
    if not ordner.is_dir():
        raise KalkulationsblattFehler(
            f"Der Ordner für die Kalkulationsblätter ist nicht erreichbar: {ordner}",
            "In der config.toml unter [pfade] den Eintrag kalkulation prüfen und sicherstellen, "
            "dass der Ordner vorhanden und lesbar ist.",
            code="kalkulation_pfad_fehlt",
        )

    befunde: list[Befund] = []
    je_projekt: dict[int, list[Path]] = {}
    for pfad in sorted(ordner.glob("*.xlsx")):
        # Excel legt beim Öffnen eine Sperrdatei '~$name.xlsx' an; sie ist keine Arbeitsmappe.
        if pfad.name.startswith("~$"):
            continue
        treffer = _DATEINAME.match(pfad.name)
        if treffer is None:
            befunde.append(
                Befund(
                    datei=pfad.name,
                    zeile=0,
                    spalte="dateiname",
                    wert=pfad.name,
                    meldung="Der Dateiname beginnt nicht mit einer Projektnummer – die Datei "
                    "wird übergangen (erwartet z. B. '26001_Mustermann.xlsx')",
                )
            )
            continue
        je_projekt.setdefault(int(treffer.group(1)), []).append(pfad)

    dateien: list[Kalkulationsdatei] = []
    for projekt_nr, pfade in sorted(je_projekt.items()):
        neueste, *weitere = sorted(pfade, key=lambda p: p.stat().st_mtime, reverse=True)
        dateien.append(Kalkulationsdatei(projekt_nr=projekt_nr, pfad=neueste))
        for uebergangen in weitere:
            befunde.append(
                Befund(
                    datei=uebergangen.name,
                    zeile=0,
                    spalte="dateiname",
                    wert=uebergangen.name,
                    meldung=f"Mehrere Kalkulationsblätter für Projekt {projekt_nr}; gelesen "
                    f"wurde die zuletzt geänderte Datei '{neueste.name}'",
                    schwere="hinweis",
                )
            )
    return dateien, befunde


# ---------------------------------------------------------------------------
# Übernehmen
# ---------------------------------------------------------------------------


@dataclass
class Uebernahmeergebnis:
    """Was ein Blatt in der Datenbank bewirkt hat."""

    projekt_nr: int | None = None
    soll_geschrieben: bool = False
    positionen_neu: int = 0
    positionen_geaendert: int = 0
    positionen_entfernt: int = 0
    positionen_behalten: int = 0
    befunde: list[Befund] = field(default_factory=list)

    @property
    def uebernommen(self) -> bool:
        return self.soll_geschrieben


def uebernehmen(sitzung: Session, blatt: Kalkulationsblatt) -> Uebernahmeergebnis:
    """Sollwerte und Stückliste eines Blattes in die Datenbank schreiben.

    Muss in einer Schreibtransaktion laufen (``schreib_transaktion``). Die Sollwerte werden je
    Projekt ersetzt – ein Kalkulationsblatt gilt als Ganzes.

    Für die Stückliste gilt der Abgleich statt des Austauschs: vorhandene Positionen werden über
    Artikelnummer bzw. Bezeichnung wiedererkannt und in ihren **Sollangaben** aktualisiert.
    ``menge_ist`` und ``bewertet_betrag`` bleiben unangetastet – sie stammen aus der Maske
    „Mengen-Ist bestätigen" und sind das Ergebnis einer Zählung, nicht einer Kalkulation.
    Positionen, die im Blatt nicht mehr vorkommen, verschwinden nur, solange nichts an ihnen
    bestätigt wurde; sonst bleiben sie stehen und werden gemeldet.
    """
    ergebnis = Uebernahmeergebnis(projekt_nr=blatt.projekt_nr, befunde=list(blatt.befunde))
    if blatt.projekt_nr is None:
        return ergebnis

    projekt = sitzung.scalar(select(Projekt).where(Projekt.projekt_nr == blatt.projekt_nr))
    if projekt is None:
        ergebnis.befunde.append(
            Befund(
                datei=blatt.datei.name,
                zeile=0,
                spalte=NAME_PROJEKT_NR,
                wert=str(blatt.projekt_nr),
                meldung=f"Es gibt kein Projekt mit der Nummer {blatt.projekt_nr} – das Blatt "
                "wird nicht übernommen",
            )
        )
        return ergebnis

    _soll_schreiben(sitzung, projekt, blatt)
    ergebnis.soll_geschrieben = True
    _stueckliste_abgleichen(sitzung, projekt, blatt, ergebnis)
    return ergebnis


def _soll_schreiben(sitzung: Session, projekt: Projekt, blatt: Kalkulationsblatt) -> None:
    soll = sitzung.get(SollKalkulation, projekt.id)
    if soll is None:
        soll = SollKalkulation(projekt_id=projekt.id)
        sitzung.add(soll)
    soll.material_soll = blatt.material_soll_cent
    soll.dl_soll = blatt.dl_soll_cent
    soll.stunden_soll = blatt.stunden_soll
    soll.marge_soll = blatt.marge_soll_promille
    soll.quelle_datei = blatt.datei.name
    soll.eingelesen_am = jetzt_utc()
    sitzung.flush()


def _stueckliste_abgleichen(
    sitzung: Session,
    projekt: Projekt,
    blatt: Kalkulationsblatt,
    ergebnis: Uebernahmeergebnis,
) -> None:
    vorhanden = list(
        sitzung.scalars(
            select(Stuecklistenposition).where(Stuecklistenposition.projekt_id == projekt.id)
        )
    )
    nach_schluessel = {_schluessel_der_zeile(z): z for z in vorhanden}
    gesehen: set[str] = set()

    for position in blatt.positionen:
        schluessel = position.schluessel
        if schluessel in gesehen:
            ergebnis.befunde.append(
                Befund(
                    datei=blatt.datei.name,
                    zeile=position.zeile,
                    spalte="artikel_nr" if position.artikel_nr else "bezeichnung",
                    wert=position.artikel_nr or position.bezeichnung,
                    meldung="Diese Position kommt im Blatt mehrfach vor; übernommen wird die "
                    "letzte Zeile",
                    schwere="hinweis",
                )
            )
        gesehen.add(schluessel)

        zeile = nach_schluessel.get(schluessel)
        if zeile is None:
            zeile = Stuecklistenposition(projekt_id=projekt.id)
            sitzung.add(zeile)
            nach_schluessel[schluessel] = zeile
            ergebnis.positionen_neu += 1
        else:
            ergebnis.positionen_geaendert += 1
        zeile.artikel_nr = position.artikel_nr
        zeile.bezeichnung = position.bezeichnung
        zeile.menge_soll = position.menge
        zeile.ek_preis = position.ek_preis_cent
        zeile.quelle = position.quelle
        zeile.gewerk = position.gewerk

    for schluessel, zeile in nach_schluessel.items():
        if schluessel in gesehen:
            continue
        if zeile.menge_ist is None and zeile.bewertet_betrag is None:
            sitzung.delete(zeile)
            ergebnis.positionen_entfernt += 1
            continue
        ergebnis.positionen_behalten += 1
        ergebnis.befunde.append(
            Befund(
                datei=blatt.datei.name,
                zeile=0,
                spalte="stueckliste",
                wert=zeile.artikel_nr or zeile.bezeichnung,
                meldung="Diese Position steht nicht mehr im Kalkulationsblatt, hat aber eine "
                "bestätigte Ist-Menge – sie bleibt erhalten und ist von Hand zu prüfen",
            )
        )
    sitzung.flush()


def _schluessel_der_zeile(zeile: Stuecklistenposition) -> str:
    if zeile.artikel_nr:
        return f"a:{zeile.artikel_nr.casefold()}"
    return f"b:{' '.join((zeile.bezeichnung or '').casefold().split())}"


# ---------------------------------------------------------------------------
# Vorlage erzeugen
# ---------------------------------------------------------------------------

# Corporate Design in der Tabelle (PLAN §11): Kopfzeile ip³ Blau mit weißer Schrift, Zebrastreifen
# in #F5F6F9, Zahlen rechtsbündig. Kein Zeichen 3 – das Wasserzeichen gehört nicht in Excel.
_BLAU = "FF2F2482"
_HELL = "FFF5F6F9"
_LINIE = "FFE0E0E0"
_SEKUNDAER = "FF666666"

VORLAGE_DATEINAME = "Kalkulationsblatt-Vorlage.xlsx"

_ZEILE_PROJEKT_NR = 6
_ZEILE_POSITIONEN_KOPF = 13
_ZEILE_POSITIONEN_START = 14

_KOPFFELDER: tuple[tuple[str, str, str, str], ...] = (
    (
        NAME_PROJEKT_NR,
        "Projektnummer",
        "0",
        "Die Nummer aus dem Leitstand, rein numerisch – zugleich der DATEV-Kostenträger.",
    ),
    (
        NAME_MATERIAL_SOLL,
        "Materialkosten Soll (netto)",
        '#,##0.00\\ "€"',
        "Kalkuliertes Material für das ganze Projekt, ohne Umsatzsteuer.",
    ),
    (
        NAME_DL_SOLL,
        "Dienstleistung Soll (netto)",
        '#,##0.00\\ "€"',
        "Fremdleistungen und Nachunternehmer, ohne Umsatzsteuer. Eigene Stunden nicht hier.",
    ),
    (
        NAME_STUNDEN_SOLL,
        "Stunden Soll",
        "#,##0.00",
        "Kalkulierte eigene Arbeitsstunden. Der Leitstand bewertet sie mit den "
        "Verrechnungssätzen aus der Konfiguration.",
    ),
    (
        NAME_MARGE_SOLL,
        "Sollmarge (%)",
        "#,##0.0",
        "Marge auf den Erlös: (Auftragswert − Kosten) ÷ Auftragswert. 18 für 18 %. "
        "Ein Prozentformat (0,18) versteht der Leitstand ebenso.",
    ),
)

_POSITIONSSPALTEN: tuple[tuple[str, int, str], ...] = (
    ("Artikelnummer", 16, "@"),
    ("Bezeichnung", 46, "@"),
    ("Menge", 12, "#,##0.000"),
    ("EK je Einheit (€)", 18, '#,##0.00\\ "€"'),
    ("Quelle", 16, "@"),
    ("Gewerk", 12, "@"),
)


def vorlage_erzeugen(ziel: Path) -> Path:
    """Erzeugt die Vorlagendatei mit dem Blatt ``EXPORT`` (PLAN §8).

    Die Datei entsteht aus Code und nicht von Hand, damit Vorlage und Einleser nicht
    auseinanderlaufen: ein Test erzeugt sie neu, liest sie und vergleicht.

    Die Positionstabelle bleibt **leer**. Eine Beispielzeile in der Vorlage wäre die erste
    Position, die jemand versehentlich mit importiert; die Beispiele stehen deshalb in der
    Hinweisspalte.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    mappe = Workbook()
    blatt = mappe.active
    blatt.title = BLATT

    weiss = Font(name="Calibri", color="FFFFFFFF", bold=True)
    blau = Font(name="Calibri", color=_BLAU, bold=True)
    grau = Font(name="Calibri", color=_SEKUNDAER, size=9)
    fuellung_blau = PatternFill("solid", fgColor=_BLAU)
    fuellung_hell = PatternFill("solid", fgColor=_HELL)
    unten = Border(bottom=Side(style="thin", color=_LINIE))

    blatt["A1"] = "ip³ Leitstand – Kalkulationsblatt, Blatt EXPORT"
    blatt["A1"].font = Font(name="Calibri", color=_BLAU, bold=True, size=14)
    blatt["A2"] = (
        "Dieses Blatt ist die Schnittstelle zum Leitstand. Die Werte in Spalte B dürfen ruhig "
        "Formeln sein, die auf die eigene Kalkulation zeigen – gelesen wird das Ergebnis."
    )
    blatt["A2"].font = grau
    blatt["A3"] = (
        "Wichtig: Die Zellen in Spalte B tragen Namen (exp_…). Bitte nicht verschieben, nicht "
        "löschen und beim Kopieren in eine andere Mappe den Namensmanager prüfen."
    )
    blatt["A3"].font = grau

    blatt["A5"] = "Kopfwerte"
    blatt["A5"].font = blau
    blatt["A5"].border = unten
    blatt["B5"].border = unten
    blatt["C5"].border = unten

    for versatz, (name, beschriftung, format_, hinweis) in enumerate(_KOPFFELDER):
        nummer = _ZEILE_PROJEKT_NR + versatz
        blatt.cell(row=nummer, column=1, value=beschriftung).font = Font(name="Calibri", bold=True)
        zelle = blatt.cell(row=nummer, column=2)
        zelle.number_format = format_
        zelle.alignment = Alignment(horizontal="right")
        zelle.fill = fuellung_hell
        blatt.cell(row=nummer, column=3, value=hinweis).font = grau
        mappe.defined_names.add(
            DefinedName(name, attr_text=f"'{BLATT}'!$B${nummer}"),
        )

    blatt.cell(row=_ZEILE_POSITIONEN_KOPF - 1, column=1, value="Stückliste").font = blau
    for spalte, (beschriftung, breite, _format) in enumerate(_POSITIONSSPALTEN, start=1):
        zelle = blatt.cell(row=_ZEILE_POSITIONEN_KOPF, column=spalte, value=beschriftung)
        zelle.font = weiss
        zelle.fill = fuellung_blau
        zelle.alignment = Alignment(horizontal="left")
        blatt.column_dimensions[get_column_letter(spalte)].width = breite

    # Formate und Zebrastreifen für die ersten Zeilen vorbereiten, damit die Liste beim Tippen
    # schon aussieht wie eine Liste. Werte stehen keine drin.
    for zeilenversatz in range(60):
        nummer = _ZEILE_POSITIONEN_START + zeilenversatz
        for spalte, (_beschriftung, _breite, format_) in enumerate(_POSITIONSSPALTEN, start=1):
            zelle = blatt.cell(row=nummer, column=spalte)
            zelle.number_format = format_
            if zeilenversatz % 2:
                zelle.fill = fuellung_hell

    mappe.defined_names.add(
        DefinedName(
            NAME_POSITIONEN_START,
            attr_text=f"'{BLATT}'!$A${_ZEILE_POSITIONEN_START}",
        ),
    )

    # Auswahllisten statt freier Eingabe: eine vertippte Quelle entscheidet darüber, ob Material
    # über DATEV oder über die Lagerbewertung ins Projekt-Ist kommt (PLAN §6.5).
    letzte = _ZEILE_POSITIONEN_START + 500
    _auswahl(
        blatt,
        STUECKLISTE_QUELLEN,
        f"E{_ZEILE_POSITIONEN_START}:E{letzte}",
        "Quelle",
        "'projektbestellt' = auf das Projekt bestellt, kommt über DATEV. "
        "'lager' = aus dem Lager entnommen, wird mit dem EK bewertet.",
    )
    _auswahl(
        blatt,
        GEWERKE_KALKULATION,
        f"F{_ZEILE_POSITIONEN_START}:F{letzte}",
        "Gewerk",
        "pv, speicher oder ls (Ladestation). Darf leer bleiben.",
    )

    blatt.column_dimensions["A"].width = max(blatt.column_dimensions["A"].width or 0, 30)
    blatt.column_dimensions["B"].width = 20
    blatt.column_dimensions["C"].width = 90
    blatt.freeze_panes = blatt.cell(row=_ZEILE_POSITIONEN_START, column=1)

    ziel.parent.mkdir(parents=True, exist_ok=True)
    mappe.save(ziel)
    mappe.close()
    return ziel


def _auswahl(blatt: Any, werte: tuple[str, ...], bereich: str, titel: str, hinweis: str) -> None:
    pruefung = DataValidation(
        type="list",
        formula1='"' + ",".join(werte) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    pruefung.promptTitle = titel
    pruefung.prompt = hinweis
    pruefung.errorTitle = titel
    pruefung.error = f"Zulässig sind: {', '.join(werte)}"
    blatt.add_data_validation(pruefung)
    pruefung.add(bereich)
