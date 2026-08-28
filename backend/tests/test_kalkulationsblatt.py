"""Vorlage und Einleser des Kalkulationsblatts (PLAN §8, EXPORT-Tab).

Die Vorlage wird in jedem Lauf neu erzeugt und wieder gelesen. Das ist der eigentliche Zweck
dieser Datei: Vorlage und Einleser dürfen nicht auseinanderlaufen, sonst liest der Leitstand
irgendwann eine Datei, die er selbst ausgegeben hat, nicht mehr.
"""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.datenbank import lese_sitzung, schreib_sitzung
from app.importe.kalkulationsblatt import (
    BLATT,
    NAME_MARGE_SOLL,
    NAME_POSITIONEN_START,
    NAME_PROJEKT_NR,
    PFLICHTNAMEN,
    VORLAGE_DATEINAME,
    ExportBlattFehlt,
    NamenFehlen,
    WerteFehlen,
    blatt_lesen,
    ordner_scannen,
    uebernehmen,
    vorlage_erzeugen,
)
from app.konfiguration import projektwurzel
from app.modelle import Projekt, SollKalkulation, Stuecklistenposition

# Zeilen und Spalten der erzeugten Vorlage – hier bewusst als Zahlen, damit ein verschobenes
# Feld im Erzeuger auffällt und nicht von derselben Konstante mitgezogen wird.
ZEILE_PROJEKT_NR = 6
ZEILE_MATERIAL = 7
ZEILE_DL = 8
ZEILE_STUNDEN = 9
ZEILE_MARGE = 10
ERSTE_POSITION = 14


def vorlage(ordner: Path, name: str = "26001_Mustermann.xlsx") -> Path:
    return vorlage_erzeugen(ordner / name)


def kopf_fuellen(
    pfad: Path,
    *,
    projekt_nr: object = 26001,
    material: object = 41500.50,
    dl: object = 8000,
    stunden: object = 120.5,
    marge: object = 18,
) -> None:
    mappe = load_workbook(pfad)
    blatt = mappe[BLATT]
    blatt.cell(row=ZEILE_PROJEKT_NR, column=2, value=projekt_nr)
    blatt.cell(row=ZEILE_MATERIAL, column=2, value=material)
    blatt.cell(row=ZEILE_DL, column=2, value=dl)
    blatt.cell(row=ZEILE_STUNDEN, column=2, value=stunden)
    blatt.cell(row=ZEILE_MARGE, column=2, value=marge)
    mappe.save(pfad)
    mappe.close()


def positionen_schreiben(pfad: Path, zeilen: list[tuple[object, ...]], ab: int = 0) -> None:
    """Zeilen der Positionstabelle setzen. ``None`` leert die Zelle.

    Bewusst über ``.value`` und nicht über ``cell(value=...)``: openpyxl übergeht dort ein
    ``None``, statt den Inhalt zu löschen – eine gestrichene Position bliebe sonst stehen.
    """
    mappe = load_workbook(pfad)
    blatt = mappe[BLATT]
    for versatz, felder in enumerate(zeilen):
        for spalte, wert in enumerate(felder, start=1):
            blatt.cell(row=ERSTE_POSITION + ab + versatz, column=spalte).value = wert
    mappe.save(pfad)
    mappe.close()


# ---------------------------------------------------------------------------
# Vorlage
# ---------------------------------------------------------------------------


def test_vorlage_traegt_alle_benannten_zellen(tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    mappe = load_workbook(pfad)
    try:
        assert BLATT in mappe.sheetnames
        for name in PFLICHTNAMEN:
            assert name in mappe.defined_names, f"Benannte Zelle {name} fehlt in der Vorlage"
    finally:
        mappe.close()


def test_vorlage_hat_keine_beispielposition(tmp_path: Path) -> None:
    """Eine Beispielzeile wäre die erste Position, die jemand versehentlich mit importiert."""
    blatt = blatt_lesen(vorlage(tmp_path))
    assert blatt.positionen == []


def test_leere_vorlage_meldet_die_fehlende_projektnummer(tmp_path: Path) -> None:
    blatt = blatt_lesen(vorlage(tmp_path))
    assert blatt.projekt_nr is None
    assert any(b.spalte == NAME_PROJEKT_NR for b in blatt.befunde)


def test_im_repo_liegende_vorlage_ist_auf_dem_stand_des_erzeugers(tmp_path: Path) -> None:
    """Die mitgelieferte Datei darf nicht hinter dem Code zurückbleiben.

    Verglichen werden die benannten Zellen und ihre Zielkoordinaten, nicht die Bytes: Excel
    schreibt bei jedem Speichern andere Zeitstempel in das Archiv.
    """
    abgelegt = projektwurzel() / "vorlagen" / VORLAGE_DATEINAME
    assert abgelegt.exists(), (
        f"{abgelegt} fehlt – mit 'ip3-leitstand kalkulationsblatt-vorlage' erzeugen"
    )
    frisch = vorlage_erzeugen(tmp_path / VORLAGE_DATEINAME)

    def namen(pfad: Path) -> dict[str, str]:
        mappe = load_workbook(pfad)
        try:
            return {name: mappe.defined_names[name].value for name in mappe.defined_names}
        finally:
            mappe.close()

    assert namen(abgelegt) == namen(frisch)


# ---------------------------------------------------------------------------
# Lesen: Kopfwerte
# ---------------------------------------------------------------------------


def test_kopfwerte_werden_in_cent_gelesen(tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    blatt = blatt_lesen(pfad)

    assert blatt.projekt_nr == 26001
    assert blatt.material_soll_cent == 4150050
    assert blatt.dl_soll_cent == 800000
    assert blatt.stunden_soll == Decimal("120.50")
    assert blatt.soll_gesamt_cent == 4950050
    assert blatt.befunde == []


@pytest.mark.parametrize(
    ("eingabe", "erwartet"),
    [
        (18, 180),  # als Zahl getippt
        (0.18, 180),  # prozentformatierte Zelle
        (Decimal("12.5"), 125),
        (1, 10),  # die Grenze: ab 1 wird als Prozent gelesen
        (0.995, 995),
        (0, 0),  # ausdrücklich getippte Null ist eine Angabe
        (None, None),  # leere Zelle: keine Sollmarge, keine Ampel
    ],
)
def test_sollmarge_wird_als_prozent_und_als_bruchteil_verstanden(
    tmp_path: Path, eingabe: object, erwartet: int | None
) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad, marge=eingabe)
    assert blatt_lesen(pfad).marge_soll_promille == erwartet


def test_unlesbare_kopfwerte_ergeben_befunde_statt_abbruch(tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad, material="ungefähr 40k", stunden="viele", marge="#DIV/0!")
    blatt = blatt_lesen(pfad)

    assert blatt.projekt_nr == 26001, "die lesbaren Werte müssen trotzdem ankommen"
    assert blatt.dl_soll_cent == 800000
    assert blatt.material_soll_cent is None
    assert blatt.stunden_soll is None
    assert blatt.marge_soll_promille is None
    assert len(blatt.befunde) == 3


def test_negatives_material_wird_nicht_uebernommen(tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad, material=-100)
    blatt = blatt_lesen(pfad)
    assert blatt.material_soll_cent is None
    assert "negativ" in blatt.befunde[0].meldung


def test_dezimalkomma_aus_einer_textzelle(tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad, material="1234,56")
    assert blatt_lesen(pfad).material_soll_cent == 123456


# ---------------------------------------------------------------------------
# Lesen: Positionen
# ---------------------------------------------------------------------------


def test_positionen_werden_gelesen_und_vereinheitlicht(tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    positionen_schreiben(
        pfad,
        [
            ("MOD-450", "Modul 450 Wp", 88, 92.40, "projektbestellt", "pv"),
            (None, "Montageschiene 4,4 m", 36, 21.90, "LAGER", "PV"),
        ],
    )
    erste, zweite = blatt_lesen(pfad).positionen

    assert erste.artikel_nr == "MOD-450"
    assert erste.menge == Decimal("88.000")
    assert erste.ek_preis_cent == 9240
    assert erste.quelle == "projektbestellt"
    assert zweite.artikel_nr is None
    assert zweite.quelle == "lager", "Groß- und Kleinschreibung darf nicht entscheiden"
    assert zweite.gewerk == "pv"


def test_eine_leerzeile_beendet_die_liste_nicht(tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    positionen_schreiben(pfad, [("A", "Erste", 1, 10, "lager", "pv")])
    positionen_schreiben(pfad, [("B", "Nach der Lücke", 2, 20, "lager", "pv")], ab=2)
    assert [p.bezeichnung for p in blatt_lesen(pfad).positionen] == ["Erste", "Nach der Lücke"]


def test_unbekannte_quelle_haelt_die_zeile_zurueck(tmp_path: Path) -> None:
    """PLAN §6.5: ohne Quelle bliebe offen, ob das Material über DATEV oder das Lager kommt."""
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    positionen_schreiben(pfad, [("X", "Unklar", 1, 10, "irgendwoher", "pv")])
    blatt = blatt_lesen(pfad)

    assert blatt.positionen == []
    assert blatt.befunde[0].spalte == "quelle"


def test_unbekanntes_gewerk_uebernimmt_die_zeile_ohne_gewerk(tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    positionen_schreiben(pfad, [("X", "Kabel", 100, 2.50, "lager", "elektrik")])
    blatt = blatt_lesen(pfad)

    assert len(blatt.positionen) == 1
    assert blatt.positionen[0].gewerk is None
    assert blatt.befunde[0].spalte == "gewerk"


def test_lagerposition_ohne_einkaufspreis_wird_gemeldet(tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    positionen_schreiben(pfad, [("X", "Kleinteile", 5, None, "lager", "pv")])
    blatt = blatt_lesen(pfad)

    assert blatt.positionen[0].ek_preis_cent is None
    assert "bewertet" in blatt.befunde[0].meldung


def test_position_ohne_menge_wird_nicht_uebernommen(tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    positionen_schreiben(pfad, [("X", "Ohne Menge", "n. B.", 10, "lager", "pv")])
    blatt = blatt_lesen(pfad)

    assert blatt.positionen == []
    assert blatt.befunde[0].spalte == "menge"


# ---------------------------------------------------------------------------
# Baufehler
# ---------------------------------------------------------------------------


def test_fehlendes_export_blatt_nennt_die_vorhandenen(tmp_path: Path) -> None:
    mappe = Workbook()
    mappe.active.title = "Kalkulation"
    pfad = tmp_path / "26001_ohne_export.xlsx"
    mappe.save(pfad)
    mappe.close()

    with pytest.raises(ExportBlattFehlt) as fehler:
        blatt_lesen(pfad)
    assert "'Kalkulation'" in fehler.value.naechster_schritt


def test_fehlende_namen_werden_alle_auf_einmal_genannt(tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    mappe = load_workbook(pfad)
    del mappe.defined_names[NAME_MARGE_SOLL]
    del mappe.defined_names[NAME_POSITIONEN_START]
    mappe.save(pfad)
    mappe.close()

    with pytest.raises(NamenFehlen) as fehler:
        blatt_lesen(pfad)
    meldung = str(fehler.value)
    assert NAME_MARGE_SOLL in meldung and NAME_POSITIONEN_START in meldung


def test_formel_ohne_gespeichertes_ergebnis_wird_erklaert(tmp_path: Path) -> None:
    """Eine Datei, die nie in Excel geöffnet wurde, trägt keine berechneten Werte."""
    pfad = vorlage(tmp_path)
    mappe = load_workbook(pfad)
    mappe[BLATT].cell(row=ZEILE_MATERIAL, column=2, value="=SUMME(Kalkulation!A1:A9)")
    mappe.save(pfad)
    mappe.close()

    with pytest.raises(WerteFehlen) as fehler:
        blatt_lesen(pfad)
    assert "in Excel öffnen" in fehler.value.naechster_schritt


def test_blattlokale_namen_werden_auch_gefunden(tmp_path: Path) -> None:
    """Excel legt Namen je nach Bedienweg in der Mappe oder im Blatt ab."""
    from openpyxl.workbook.defined_name import DefinedName

    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    mappe = load_workbook(pfad)
    del mappe.defined_names[NAME_PROJEKT_NR]
    mappe[BLATT].defined_names.add(
        DefinedName(NAME_PROJEKT_NR, attr_text=f"'{BLATT}'!$B${ZEILE_PROJEKT_NR}")
    )
    mappe.save(pfad)
    mappe.close()

    assert blatt_lesen(pfad).projekt_nr == 26001


# ---------------------------------------------------------------------------
# Ordner scannen
# ---------------------------------------------------------------------------


def test_ordner_scannen_liest_die_projektnummer_aus_dem_dateinamen(tmp_path: Path) -> None:
    vorlage(tmp_path, "26001_Mustermann.xlsx")
    vorlage(tmp_path, "26002 Zweitkunde.xlsx")
    vorlage(tmp_path, "Angebot_ohne_Nummer.xlsx")
    (tmp_path / "~$26001_Mustermann.xlsx").write_bytes(b"Sperrdatei")

    dateien, befunde = ordner_scannen(tmp_path)

    assert [d.projekt_nr for d in dateien] == [26001, 26002]
    assert len(befunde) == 1
    assert befunde[0].datei == "Angebot_ohne_Nummer.xlsx"


def test_mehrere_blaetter_je_projekt_nehmen_das_neueste(tmp_path: Path) -> None:
    import os
    import time

    alt = vorlage(tmp_path, "26001_alt.xlsx")
    neu = vorlage(tmp_path, "26001_neu.xlsx")
    frueher = time.time() - 3600
    os.utime(alt, (frueher, frueher))

    dateien, befunde = ordner_scannen(tmp_path)

    assert [d.pfad for d in dateien] == [neu]
    assert befunde[0].schwere == "hinweis"
    assert "26001_neu.xlsx" in befunde[0].meldung


# ---------------------------------------------------------------------------
# Übernehmen
# ---------------------------------------------------------------------------


@pytest.fixture
def projekt(gesäte_db) -> int:
    """Ein Projekt 26001, an dem die Übernahme hängt. Liefert die Datenbank-ID."""
    from app.modelle import Firma, Kunde

    with schreib_sitzung() as sitzung:
        firma_id = sitzung.scalar(select(Firma.id).order_by(Firma.id).limit(1))
        kunde = Kunde(kunden_nr=14001, name="Kalkulation GmbH", ort="Weiden", typ="b2b")
        sitzung.add(kunde)
        sitzung.flush()
        eintrag = Projekt(
            projekt_nr=26001,
            firma_id=firma_id,
            kunde_id=kunde.id,
            status="in_bau",
            ab_wert_netto=6000000,
        )
        sitzung.add(eintrag)
        sitzung.flush()
        return eintrag.id


def stueckliste(sitzung) -> list[Stuecklistenposition]:
    return list(sitzung.scalars(select(Stuecklistenposition).order_by(Stuecklistenposition.id)))


def test_uebernahme_schreibt_soll_und_stueckliste(projekt: int, tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    positionen_schreiben(
        pfad,
        [
            ("MOD-450", "Modul 450 Wp", 88, 92.40, "projektbestellt", "pv"),
            ("SCH-44", "Montageschiene", 36, 21.90, "lager", "pv"),
        ],
    )
    blatt = blatt_lesen(pfad)
    with schreib_sitzung() as sitzung:
        ergebnis = uebernehmen(sitzung, blatt)

    assert ergebnis.soll_geschrieben and ergebnis.positionen_neu == 2
    with lese_sitzung() as sitzung:
        soll = sitzung.get(SollKalkulation, projekt)
        assert soll.material_soll == 4150050
        assert soll.dl_soll == 800000
        assert soll.marge_soll == 180
        assert soll.quelle_datei == "26001_Mustermann.xlsx"
        assert soll.eingelesen_am is not None
        assert len(stueckliste(sitzung)) == 2


def test_zweites_einlesen_ersetzt_die_sollwerte(projekt: int, tmp_path: Path) -> None:
    """Ein Kalkulationsblatt gilt als Ganzes; es gibt je Projekt einen Satz Sollwerte."""
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    with schreib_sitzung() as sitzung:
        uebernehmen(sitzung, blatt_lesen(pfad))

    kopf_fuellen(pfad, material=50000, marge=15)
    with schreib_sitzung() as sitzung:
        uebernehmen(sitzung, blatt_lesen(pfad))

    with lese_sitzung() as sitzung:
        assert len(list(sitzung.scalars(select(SollKalkulation)))) == 1
        soll = sitzung.get(SollKalkulation, projekt)
        assert soll.material_soll == 5000000
        assert soll.marge_soll == 150


def test_unbekannte_projektnummer_wird_gemeldet_statt_angelegt(
    projekt: int, tmp_path: Path
) -> None:
    pfad = vorlage(tmp_path, "99999_Unbekannt.xlsx")
    kopf_fuellen(pfad, projekt_nr=99999)
    with schreib_sitzung() as sitzung:
        ergebnis = uebernehmen(sitzung, blatt_lesen(pfad))

    assert not ergebnis.soll_geschrieben
    assert "kein Projekt" in ergebnis.befunde[0].meldung
    with lese_sitzung() as sitzung:
        assert list(sitzung.scalars(select(SollKalkulation))) == []


def test_erneutes_einlesen_erhaelt_die_bestaetigte_menge(projekt: int, tmp_path: Path) -> None:
    """Das Blatt ist die Wahrheit für das Soll, die Bestätigungsmaske für das Ist."""
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    positionen_schreiben(pfad, [("MOD-450", "Modul 450 Wp", 88, 92.40, "lager", "pv")])
    with schreib_sitzung() as sitzung:
        uebernehmen(sitzung, blatt_lesen(pfad))
    with schreib_sitzung() as sitzung:
        position = stueckliste(sitzung)[0]
        position.menge_ist = Decimal("86.000")
        position.bewertet_betrag = 794640

    positionen_schreiben(pfad, [("MOD-450", "Modul 450 Wp", 90, 95.00, "lager", "pv")])
    with schreib_sitzung() as sitzung:
        ergebnis = uebernehmen(sitzung, blatt_lesen(pfad))

    assert ergebnis.positionen_geaendert == 1 and ergebnis.positionen_neu == 0
    with lese_sitzung() as sitzung:
        position = stueckliste(sitzung)[0]
        assert position.menge_soll == Decimal("90.000"), "das Soll folgt dem Blatt"
        assert position.ek_preis == 9500
        assert position.menge_ist == Decimal("86.000"), "die gezählte Menge bleibt"
        assert position.bewertet_betrag == 794640


def test_entfallene_position_ohne_bestaetigung_verschwindet(projekt: int, tmp_path: Path) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    positionen_schreiben(
        pfad,
        [
            ("A-1", "Bleibt", 1, 10, "lager", "pv"),
            ("A-2", "Entfällt", 2, 20, "lager", "pv"),
        ],
    )
    with schreib_sitzung() as sitzung:
        uebernehmen(sitzung, blatt_lesen(pfad))

    positionen_schreiben(pfad, [("A-1", "Bleibt", 1, 10, "lager", "pv"), (None,) * 6])
    with schreib_sitzung() as sitzung:
        ergebnis = uebernehmen(sitzung, blatt_lesen(pfad))

    assert ergebnis.positionen_entfernt == 1
    with lese_sitzung() as sitzung:
        assert [p.artikel_nr for p in stueckliste(sitzung)] == ["A-1"]


def test_entfallene_position_mit_bestaetigung_bleibt_und_wird_gemeldet(
    projekt: int, tmp_path: Path
) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    positionen_schreiben(pfad, [("A-2", "Entfällt", 2, 20, "lager", "pv")])
    with schreib_sitzung() as sitzung:
        uebernehmen(sitzung, blatt_lesen(pfad))
    with schreib_sitzung() as sitzung:
        stueckliste(sitzung)[0].menge_ist = Decimal("2.000")

    positionen_schreiben(pfad, [(None,) * 6])
    with schreib_sitzung() as sitzung:
        ergebnis = uebernehmen(sitzung, blatt_lesen(pfad))

    assert ergebnis.positionen_behalten == 1 and ergebnis.positionen_entfernt == 0
    assert "bestätigte Ist-Menge" in ergebnis.befunde[0].meldung
    with lese_sitzung() as sitzung:
        assert len(stueckliste(sitzung)) == 1


def test_position_ohne_artikelnummer_wird_ueber_die_bezeichnung_wiedererkannt(
    projekt: int, tmp_path: Path
) -> None:
    pfad = vorlage(tmp_path)
    kopf_fuellen(pfad)
    positionen_schreiben(pfad, [(None, "Montageschiene 4,4 m", 36, 21.90, "lager", "pv")])
    with schreib_sitzung() as sitzung:
        uebernehmen(sitzung, blatt_lesen(pfad))

    positionen_schreiben(pfad, [(None, "Montageschiene 4,4 m", 40, 22.50, "lager", "pv")])
    with schreib_sitzung() as sitzung:
        ergebnis = uebernehmen(sitzung, blatt_lesen(pfad))

    assert ergebnis.positionen_neu == 0 and ergebnis.positionen_geaendert == 1
    with lese_sitzung() as sitzung:
        assert stueckliste(sitzung)[0].menge_soll == Decimal("40.000")
