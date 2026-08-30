"""Angebotsliste aus dem Angebots-Tool einlesen (PLAN §7 Phase 7).

**Die echte Datei liegt noch nicht vor.** Deshalb ist dieser Leser so gebaut, dass er sich an
sie anpassen lässt, ohne dass Code geändert wird: Spaltennamen und die Übersetzung der
Statuswerte stehen in der ``config.toml`` unter ``[angebote]``, genau wie bei DATEV und TimeTac.
Gelesen werden Excel-Mappen (``.xlsx``) und Textdateien (``.csv``) – welches Format das Tool
ausgibt, entscheidet sich erst, wenn es da ist, und beides zu können kostet wenig.

Zwei Regeln, die den Import ungefährlich machen:

* **Wiedererkennen statt anhäufen.** Eine Zeile mit bekannter Angebotsnummer aktualisiert das
  vorhandene Angebot. Ohne Nummer entsteht ein neues – dann ist ein zweiter Lauf derselben
  Datei aber ein Duplikat, und genau das sagt der Befund.
* **Nichts still verwerfen.** Eine unlesbare Summe, ein unbekannter Status, ein Monat, der
  keiner ist: jede solche Zeile wird übernommen, soweit es geht, und erscheint als Befund im
  Importprotokoll (PLAN §2).

Ein bereits **gewonnenes** Angebot wird nicht mehr angefasst. Daran hängt ein Projekt; eine
Statusänderung aus der Quelldatei würde den Auftragsbestand verändern, und das darf ein Import
nicht (CLAUDE.md Regel 5 sinngemäß).
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.fehler import FachFehler
from app.importe.befunde import Befund
from app.importe.csv_leser import CsvDatei, aus_zeilen, deutsche_zahl, deutsches_datum
from app.importe.csv_leser import lesen as csv_lesen
from app.modelle import Angebot, Kunde
from app.zeit import jetzt_utc, monat_gueltig

# Ohne Kunde und Summe ist eine Zeile kein Angebot.
PFLICHTFELDER: tuple[str, ...] = ("kunde", "summe")

EXCEL_ENDUNGEN = (".xlsx", ".xlsm")


class AngebotsdateiFehlt(FachFehler):
    code = "angebotsdatei_fehlt"

    def __init__(self, pfad: Path) -> None:
        super().__init__(
            f"Die Angebotsdatei '{pfad}' gibt es nicht.",
            "Bitte den Pfad prüfen. Die Datei wird nur gelesen und nicht verändert.",
        )


@dataclass
class Angebotszeile:
    """Eine gelesene Zeile, schon in die Form des Leitstands gebracht."""

    zeile: int
    angebot_nr: str | None
    kunde_name: str
    bezeichnung: str | None
    summe_cent: int
    wahrscheinlichkeit_promille: int
    erwarteter_monat: str | None
    status: str
    datum: date | None


@dataclass
class Angebotsdatei:
    pfad: Path
    zeilen: list[Angebotszeile] = field(default_factory=list)
    befunde: list[Befund] = field(default_factory=list)


@dataclass
class Uebernahme:
    """Was der Import geschrieben hat."""

    neu: int = 0
    aktualisiert: int = 0
    uebersprungen: int = 0
    befunde: list[Befund] = field(default_factory=list)


def _excel_zeilen(pfad: Path) -> tuple[list[str], list[list[str]]]:
    """Erste Tabelle einer Excel-Mappe als Kopfzeile und Datenzeilen, alles als Text.

    ``data_only=True``: das Angebots-Tool rechnet mit Formeln, und ein Import interessiert sich
    für Ergebnisse, nicht für Rechenwege. Damit gilt allerdings, was Excel zuletzt gespeichert
    hat – eine Mappe, die nie in Excel geöffnet wurde, hat dort leere Zellen.
    """
    mappe = load_workbook(pfad, data_only=True, read_only=True)
    try:
        blatt = mappe[mappe.sheetnames[0]]
        zeilen = [
            ["" if zelle is None else str(zelle).strip() for zelle in reihe]
            for reihe in blatt.iter_rows(values_only=True)
        ]
    finally:
        mappe.close()

    # Kopfzeile ist die erste Zeile mit mindestens zwei gefüllten Zellen: Angebotslisten tragen
    # oft eine Überschrift oder eine Leerzeile darüber.
    for index, reihe in enumerate(zeilen):
        if sum(1 for zelle in reihe if zelle) >= 2:
            return reihe, zeilen[index + 1 :]
    return ([], [])


def datei_lesen(pfad: Path, zuordnung: dict[str, list[str]]) -> CsvDatei:
    """Angebotsdatei als Tabelle mit den Feldnamen des Leitstands."""
    if not pfad.exists():
        raise AngebotsdateiFehlt(pfad)
    if pfad.suffix.lower() in EXCEL_ENDUNGEN:
        kopf, datenzeilen = _excel_zeilen(pfad)
        return aus_zeilen(pfad, kopf, datenzeilen, zuordnung, pflicht=PFLICHTFELDER)
    return csv_lesen(pfad, zuordnung, pflicht=PFLICHTFELDER)


def _wahrscheinlichkeit(inhalt: str) -> int | None:
    """``'60 %'``, ``'60'``, ``'0,6'`` – alles wird zu Promille.

    Werte bis 1 gelten als Anteil (0,6 = 60 %), darüber als Prozent. Die Grenze ist eine
    Annahme, aber die einzige, die beide üblichen Schreibweisen richtig deutet: „1" als ein
    Prozent zu lesen wäre bei einer Angebotsliste absurder, als es als 100 % zu lesen.
    """
    roh = inhalt.replace("%", "").strip()
    zahl = deutsche_zahl(roh)
    if zahl is None:
        return None
    anteil = zahl if zahl > 1 else zahl * 100
    # Ungekürzt zurück: ob 150 % ein Tippfehler ist, entscheidet der Aufrufer – und meldet es.
    return int((anteil * 10).to_integral_value())


def _monat(inhalt: str) -> str | None:
    """Erwarteter Auftragsmonat: ``'2026-09'``, ``'09/2026'``, ``'09.2026'`` oder ein Datum."""
    roh = inhalt.strip()
    if not roh:
        return None
    if monat_gueltig(roh):
        return roh
    for trenner in ("/", ".", "-"):
        if trenner in roh:
            teile = [t.strip() for t in roh.split(trenner)]
            if len(teile) == 2 and all(t.isdigit() for t in teile):
                monat, jahr = teile
                if len(monat) == 4:  # '2026-09'
                    monat, jahr = jahr, monat
                if len(jahr) == 2:
                    jahr = f"20{jahr}"
                if len(jahr) == 4 and 1 <= int(monat) <= 12:
                    return f"{jahr}-{int(monat):02d}"
    tag = deutsches_datum(roh)
    return f"{tag:%Y-%m}" if tag else None


def lesen(
    pfad: Path,
    zuordnung: dict[str, list[str]],
    status_zuordnung: dict[str, str],
    *,
    standard_wahrscheinlichkeit: int,
) -> Angebotsdatei:
    """Angebotsdatei einlesen. Unklare Werte werden zu Befunden, nicht zu Ausfällen."""
    tabelle = datei_lesen(pfad, zuordnung)
    ergebnis = Angebotsdatei(pfad=pfad)
    uebersetzung = {schluessel.casefold(): wert for schluessel, wert in status_zuordnung.items()}

    for zeile in tabelle.zeilen:
        kunde_name = zeile.wert("kunde").strip()
        if not kunde_name:
            ergebnis.befunde.append(
                Befund(pfad.name, zeile.nummer, "kunde", "", "Zeile ohne Kunde, übergangen")
            )
            continue

        summe = deutsche_zahl(zeile.wert("summe"))
        if summe is None:
            ergebnis.befunde.append(
                Befund(
                    pfad.name,
                    zeile.nummer,
                    "summe",
                    zeile.wert("summe"),
                    "Angebotssumme nicht lesbar, Zeile übergangen",
                )
            )
            continue
        if summe < 0:
            ergebnis.befunde.append(
                Befund(
                    pfad.name,
                    zeile.nummer,
                    "summe",
                    zeile.wert("summe"),
                    "Negative Angebotssumme, als Betrag ohne Vorzeichen übernommen",
                )
            )
            summe = abs(summe)

        wahrscheinlichkeit = _wahrscheinlichkeit(zeile.wert("wahrscheinlichkeit"))
        if wahrscheinlichkeit is None:
            if zeile.wert("wahrscheinlichkeit").strip():
                ergebnis.befunde.append(
                    Befund(
                        pfad.name,
                        zeile.nummer,
                        "wahrscheinlichkeit",
                        zeile.wert("wahrscheinlichkeit"),
                        f"Wahrscheinlichkeit nicht lesbar, {standard_wahrscheinlichkeit / 10:.0f} "
                        "% angenommen",
                    )
                )
            wahrscheinlichkeit = standard_wahrscheinlichkeit
        elif not 0 <= wahrscheinlichkeit <= 1000:
            ergebnis.befunde.append(
                Befund(
                    pfad.name,
                    zeile.nummer,
                    "wahrscheinlichkeit",
                    zeile.wert("wahrscheinlichkeit"),
                    "Wahrscheinlichkeit außerhalb von 0 bis 100 %, auf den Rand gesetzt",
                )
            )
            wahrscheinlichkeit = max(0, min(1000, wahrscheinlichkeit))

        roh_monat = zeile.wert("erwarteter_monat")
        monat = _monat(roh_monat)
        if monat is None and roh_monat.strip():
            ergebnis.befunde.append(
                Befund(
                    pfad.name,
                    zeile.nummer,
                    "erwarteter_monat",
                    roh_monat,
                    "Erwarteter Monat nicht lesbar, Angebot bleibt unterminiert",
                )
            )

        roh_status = zeile.wert("status").strip()
        status = uebersetzung.get(roh_status.casefold(), "offen")
        if roh_status and roh_status.casefold() not in uebersetzung:
            ergebnis.befunde.append(
                Befund(
                    pfad.name,
                    zeile.nummer,
                    "status",
                    roh_status,
                    "Unbekannter Status, als 'offen' übernommen. Ergänzbar unter "
                    "[angebote.status_zuordnung] in der config.toml",
                    schwere="hinweis",
                )
            )

        ergebnis.zeilen.append(
            Angebotszeile(
                zeile=zeile.nummer,
                angebot_nr=(zeile.wert("angebot_nr").strip() or None),
                kunde_name=kunde_name,
                bezeichnung=(zeile.wert("bezeichnung").strip() or None),
                summe_cent=int((summe * 100).to_integral_value()),
                wahrscheinlichkeit_promille=wahrscheinlichkeit,
                erwarteter_monat=monat,
                status=status,
                datum=deutsches_datum(zeile.wert("datum")),
            )
        )

    if not tabelle.hat("angebot_nr"):
        ergebnis.befunde.append(
            Befund(
                pfad.name,
                0,
                "angebot_nr",
                "",
                "Die Datei führt keine Angebotsnummer. Ein zweiter Lauf legt die Angebote "
                "deshalb erneut an, statt sie zu aktualisieren",
                schwere="hinweis",
            )
        )
    return ergebnis


def uebernehmen(sitzung: Session, datei: Angebotsdatei) -> Uebernahme:
    """Gelesene Angebote schreiben. Muss in einer Schreibtransaktion laufen."""
    ergebnis = Uebernahme(befunde=list(datei.befunde))
    jetzt = jetzt_utc()

    for zeile in datei.zeilen:
        vorhanden = (
            sitzung.scalar(select(Angebot).where(Angebot.angebot_nr == zeile.angebot_nr))
            if zeile.angebot_nr
            else None
        )

        if vorhanden is not None and vorhanden.status == "gewonnen":
            # Daran hängt ein Projekt. Ein Import darf den Auftragsbestand nicht verändern.
            ergebnis.uebersprungen += 1
            ergebnis.befunde.append(
                Befund(
                    datei.pfad.name,
                    zeile.zeile,
                    "status",
                    zeile.status,
                    f"Angebot {zeile.angebot_nr} ist bereits gewonnen und wurde nicht "
                    "überschrieben",
                    schwere="hinweis",
                )
            )
            continue

        if vorhanden is None:
            vorhanden = Angebot(angebot_nr=zeile.angebot_nr, kunde_name=zeile.kunde_name)
            sitzung.add(vorhanden)
            ergebnis.neu += 1
        else:
            ergebnis.aktualisiert += 1

        vorhanden.kunde_name = zeile.kunde_name
        vorhanden.bezeichnung = zeile.bezeichnung
        vorhanden.summe_netto = zeile.summe_cent
        vorhanden.wahrscheinlichkeit_promille = zeile.wahrscheinlichkeit_promille
        vorhanden.erwarteter_monat = zeile.erwarteter_monat
        vorhanden.status = zeile.status
        vorhanden.datum = zeile.datum
        vorhanden.quelle_datei = datei.pfad.name
        vorhanden.eingelesen_am = jetzt
        if vorhanden.kunde_id is None:
            vorhanden.kunde_id = _kunde_finden(sitzung, zeile.kunde_name)

    sitzung.flush()
    return ergebnis


def _kunde_finden(sitzung: Session, name: str) -> int | None:
    """Kunden über den Namen zuordnen – exakt, ohne Ähnlichkeitssuche.

    Ein Angebot geht oft an einen Interessenten, der noch kein Kunde ist; dann bleibt die
    Zuordnung leer und der Name steht als Text daneben. Eine unscharfe Zuordnung wäre hier
    gefährlicher als keine: sie hängte ein Angebot an die falsche Firma.
    """
    return sitzung.scalar(select(Kunde.id).where(Kunde.name == name.strip()))
